"""F14: conversión, impresión, orden y export de métricas + fugas P0.

- §0a: el XLSX del vendedor NO trae las columnas "Tipo de cambio" ni
  "Confirmado MXN" (ni encabezado ni celda); el comprador sí.
- §0b: dinero_confirmado / _desglose / _mxn NO EXISTEN en el JSON del
  vendedor en /resumen, /por-sucursal, /por-cliente y /serie (patrón
  proveedor, §4.9); comprador, gerente y admin sí los reciben.
- p.1: conversión por CICLOS del periodo con ARITMÉTICA A MANO, exclusiones
  (todo-no-encontrado, canceladas antes de cotizar) y denominador cero.
- p.2: bitácora de impresiones — documento por estatus, reimpresión de la
  cotización tras confirmar, snapshot sin FKs y 422 en estados previos.
- p.4: export de comparativas — tipos REALES de celda, orden numérico con
  vacíos al final, criterio V/A/R por rojas, gates por dimensión y columna
  de dinero ausente para el vendedor.
"""

from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.models.cotizacion import CotizacionOpcion, Letra, Moneda, OpcionPartida
from app.models.historial import HistorialEstado
from app.models.impresion import Impresion
from app.models.solicitud import Estado, Prioridad, Solicitud, SolicitudPartida
from app.models.sucursal import CompradorSucursal
from app.models.usuario import Rol

BASE = "/api/v1/solicitudes"
METRICAS = "/api/v1/metricas"
EXPORT = "/api/v1/solicitudes/export"
EXPORT_COMP = "/api/v1/metricas/export-comparativas"

PARTIDA = {"cantidad": "10", "unidad": "PZ", "descripcion": "SOLERA 1/8 X 1"}
HOY = {"desde": "2026-01-01", "hasta": "2030-12-31"}


@pytest.fixture
def entorno(db, make_user, make_sucursal):
    sucursal = make_sucursal("F14 Suc")
    comprador = make_user(Rol.COMPRADOR)
    db.add(CompradorSucursal(comprador_id=comprador.id, sucursal_id=sucursal.id, titular=True))
    db.commit()
    return SimpleNamespace(
        sucursal=sucursal,
        comprador=comprador,
        vendedor=make_user(Rol.VENDEDOR, sucursal_id=sucursal.id),
        ajeno=make_user(Rol.VENDEDOR, sucursal_id=sucursal.id),
        gerente=make_user(Rol.GERENTE_SUCURSAL, sucursal_id=sucursal.id),
        gerente_compras=make_user(Rol.GERENTE_COMPRAS),
        director=make_user(Rol.DIRECTOR_VENTAS),
        admin=make_user(Rol.ADMIN),
    )


def _enviada(client, entorno, auth_headers, partidas=None):
    headers = auth_headers(entorno.vendedor)
    r = client.post(
        BASE, headers=headers, json={"cliente": "DINCO", "partidas": partidas or [PARTIDA]}
    )
    assert r.status_code == 201, r.text
    sid = r.json()["id"]
    assert client.post(f"{BASE}/{sid}/enviar", headers=headers).status_code == 200
    return sid


def _capturar_a(client, entorno, auth_headers, sid, renglones_de=None):
    headers = auth_headers(entorno.comprador)
    detalle = client.get(f"{BASE}/{sid}", headers=headers).json()
    if renglones_de is None:
        renglones_de = lambda p: {  # noqa: E731
            "partida_id": p["id"],
            "moneda": "MXN",
            "precio_unitario": "100.00",
            "tiempo_entrega": "1 semana",
        }
    renglones = [renglones_de(p) for p in detalle["partidas"]]
    r = client.put(
        f"{BASE}/{sid}/opciones/A",
        headers=headers,
        json={"vigencia": "2026-12-31", "renglones": renglones},
    )
    assert r.status_code == 200, r.text


def _cotizada(client, entorno, auth_headers, partidas=None, renglones_de=None):
    sid = _enviada(client, entorno, auth_headers, partidas)
    _capturar_a(client, entorno, auth_headers, sid, renglones_de)
    r = client.post(f"{BASE}/{sid}/cotizar", headers=auth_headers(entorno.comprador), json={})
    assert r.status_code == 200, r.text
    return sid


def _confirmada(client, entorno, auth_headers, con_comprobante):
    sid = _cotizada(client, entorno, auth_headers)
    con_comprobante(sid, entorno.vendedor)
    r = client.post(
        f"{BASE}/{sid}/seleccionar", headers=auth_headers(entorno.vendedor), json={"letra": "A"}
    )
    assert r.status_code == 200, r.text
    return sid


_folio = iter(range(1, 10_000))


def _sintetica(db, entorno, *, estado, eventos=(), **campos):
    """Solicitud con historial SINTÉTICO de timestamps controlados (mismo
    patrón permitido en test_metricas) para la aritmética a mano."""
    solicitud = Solicitud(
        folio=f"F14-{next(_folio)}",
        vendedor_id=entorno.vendedor.id,
        sucursal_id=entorno.sucursal.id,
        comprador_id=entorno.comprador.id,
        estado=estado,
        prioridad=Prioridad.NORMAL,
        **campos,
    )
    db.add(solicitud)
    db.flush()
    for de, a, ts in eventos:
        db.add(
            HistorialEstado(
                solicitud_id=solicitud.id, de=de, a=a, usuario_id=entorno.vendedor.id, timestamp=ts
            )
        )
    db.commit()
    return solicitud


def _utc(y, m, d, hh=12):
    return datetime(y, m, d, hh, tzinfo=UTC)


# ------------------------------------------------------------ §0a export XLSX


def _hoja(respuesta):
    assert respuesta.status_code == 200, respuesta.text
    return load_workbook(BytesIO(respuesta.content)).active


def test_export_vendedor_sin_tc_ni_consolidado(client, db, entorno, auth_headers, con_comprobante):
    """§0a: vendedor exporta una CONFIRMADA en USD → el XLSX NO trae las
    columnas de TC/consolidado; el comprador sí (con sus valores)."""
    sid = _enviada(client, entorno, auth_headers)
    _capturar_a(
        client,
        entorno,
        auth_headers,
        sid,
        lambda p: {
            "partida_id": p["id"],
            "moneda": "USD",
            "precio_unitario": "50.00",
            "tiempo_entrega": "2 semanas",
        },
    )
    r = client.post(
        f"{BASE}/{sid}/cotizar",
        headers=auth_headers(entorno.comprador),
        json={"tipo_cambio": "18.5000"},
    )
    assert r.status_code == 200, r.text
    con_comprobante(sid, entorno.vendedor)
    r = client.post(
        f"{BASE}/{sid}/seleccionar", headers=auth_headers(entorno.vendedor), json={"letra": "A"}
    )
    assert r.status_code == 200, r.text

    ws = _hoja(client.get(EXPORT, headers=auth_headers(entorno.vendedor)))
    encabezados = [c.value for c in ws[1]]
    assert "Tipo de cambio" not in encabezados
    assert "Confirmado MXN" not in encabezados
    # El resto de columnas sigue completo y alineado (celdas = encabezados).
    assert "Monto USD" in encabezados
    fila = next(f for f in ws.iter_rows(min_row=2, values_only=True))
    assert len(fila) == len(encabezados)
    col_usd = encabezados.index("Monto USD")
    assert fila[col_usd] == 500  # 10 × 50.00 USD, aritmética a mano

    ws = _hoja(client.get(EXPORT, headers=auth_headers(entorno.comprador)))
    encabezados = [c.value for c in ws[1]]
    assert "Tipo de cambio" in encabezados and "Confirmado MXN" in encabezados
    fila = next(f for f in ws.iter_rows(min_row=2, values_only=True))
    assert fila[encabezados.index("Tipo de cambio")] == 18.5
    # 500 USD × 18.5 = 9,250.00 MXN consolidado — a mano.
    assert fila[encabezados.index("Confirmado MXN")] == 9250


# ------------------------------------------------ §0b claves por rol en JSON

_CLAVES_RESUMEN = ("dinero_confirmado", "dinero_confirmado_desglose")


def test_resumen_vendedor_sin_claves_de_consolidado(client, entorno, auth_headers):
    _cotizada(client, entorno, auth_headers)
    body = client.get(f"{METRICAS}/resumen", headers=auth_headers(entorno.vendedor)).json()
    for clave in _CLAVES_RESUMEN:
        assert clave not in body  # la clave NO EXISTE, no viene vacía
    assert "dinero_referencia" in body  # la referencia sí es suya (§4.9)
    # Comprador, gerente de sucursal y admin SÍ reciben el consolidado.
    for u in (entorno.comprador, entorno.gerente, entorno.admin):
        body = client.get(f"{METRICAS}/resumen", headers=auth_headers(u)).json()
        for clave in _CLAVES_RESUMEN:
            assert clave in body, u.rol


@pytest.mark.parametrize("endpoint", ["por-sucursal", "por-cliente"])
def test_tablas_vendedor_sin_dinero_confirmado(client, entorno, auth_headers, endpoint):
    _cotizada(client, entorno, auth_headers)
    filas = client.get(
        f"{METRICAS}/{endpoint}", params=HOY, headers=auth_headers(entorno.vendedor)
    ).json()
    assert filas, "el vendedor debe ver su propia fila"
    assert all("dinero_confirmado" not in fila for fila in filas)
    filas = client.get(
        f"{METRICAS}/{endpoint}", params=HOY, headers=auth_headers(entorno.admin)
    ).json()
    assert filas and all("dinero_confirmado" in fila for fila in filas)


def test_serie_vendedor_sin_dinero_confirmado_mxn(client, entorno, auth_headers):
    _cotizada(client, entorno, auth_headers)
    hoy = datetime.now(UTC).date().isoformat()
    params = {"desde": hoy, "hasta": hoy}
    semanas = client.get(
        f"{METRICAS}/serie", params=params, headers=auth_headers(entorno.vendedor)
    ).json()["semanas"]
    assert semanas and all("dinero_confirmado_mxn" not in s for s in semanas)
    semanas = client.get(
        f"{METRICAS}/serie", params=params, headers=auth_headers(entorno.comprador)
    ).json()["semanas"]
    assert semanas and all("dinero_confirmado_mxn" in s for s in semanas)


# ------------------------------------------------------------ p.1 conversión


def test_conversion_aritmetica_a_mano(client, db, entorno, auth_headers, con_comprobante):
    """Flujo REAL por API. A mano: cotizadas = {S1, S2, S5} = 3 (S1 recotizada
    dos veces cuenta UNA; S4 todo-no-encontrado y S6 cancelada-sin-cotizar se
    excluyen); confirmadas = {S2} = 1 → tasa = 1/3 = 0.3333."""
    s1 = _cotizada(client, entorno, auth_headers)
    # S1 "recotizada": segunda transición REAL →COTIZADA (reversión de una
    # NO_CONFIRMADA por el admin) — el denominador la cuenta UNA vez.
    db.add(
        HistorialEstado(
            solicitud_id=s1,
            de=Estado.NO_CONFIRMADA,
            a=Estado.COTIZADA,
            usuario_id=entorno.admin.id,
            timestamp=datetime.now(UTC),
        )
    )
    db.commit()
    _confirmada(client, entorno, auth_headers, con_comprobante)  # S2
    # S4: cotizada con TODOS los renglones no encontrados → excluida. El flujo
    # actual lo IMPIDE por API (cotizar exige ≥1 renglón cotizado y pide
    # rechazar), así que se inyecta sintética: fija la exclusión para datos
    # legados y como guardia si esa validación cambiara.
    s4 = _sintetica(
        db,
        entorno,
        estado=Estado.COTIZADA,
        cotizado_en=datetime.now(UTC),
        eventos=[(Estado.EN_PROCESO, Estado.COTIZADA, datetime.now(UTC))],
    )
    partida = SolicitudPartida(
        solicitud_id=s4.id, num_partida=1, cantidad=1, unidad="PZ", descripcion="NO HAY"
    )
    db.add(partida)
    db.flush()
    opcion = CotizacionOpcion(solicitud_id=s4.id, letra=Letra.A, completa=True)
    db.add(opcion)
    db.flush()
    db.add(
        OpcionPartida(
            opcion_id=opcion.id,
            partida_id=partida.id,
            cantidad=1,
            unidad="PZ",
            no_encontrada=True,
        )
    )
    db.commit()
    # S5: mixta — un renglón no encontrado y uno cotizado → SÍ cuenta.
    s5 = _cotizada(
        client,
        entorno,
        auth_headers,
        partidas=[PARTIDA, {"cantidad": "5", "unidad": "KG", "descripcion": "PLACA 1/4"}],
        renglones_de=lambda p: (
            {"partida_id": p["id"], "no_encontrada": True}
            if p["num_partida"] == 1
            else {
                "partida_id": p["id"],
                "moneda": "MXN",
                "precio_unitario": "10.00",
                "tiempo_entrega": "1 semana",
            }
        ),
    )
    assert s5
    # S6: cancelada ANTES de cotizar → jamás entra al denominador.
    s6 = _enviada(client, entorno, auth_headers)
    assert (
        client.post(f"{BASE}/{s6}/cancelar", headers=auth_headers(entorno.vendedor)).status_code
        == 200
    )

    conv = client.get(f"{METRICAS}/resumen", headers=auth_headers(entorno.admin)).json()[
        "conversion"
    ]
    assert conv["cotizadas"] == 3
    assert conv["confirmadas"] == 1
    assert conv["tasa"] == round(1 / 3, 4) == 0.3333


def test_conversion_denominador_cero(client, entorno, auth_headers):
    """Sin cotizadas en el periodo → tasa None (la UI muestra '—')."""
    _enviada(client, entorno, auth_headers)  # enviada, aún sin cotizar
    body = client.get(f"{METRICAS}/resumen", headers=auth_headers(entorno.admin)).json()
    conv = body["conversion"]
    assert conv == {
        "cotizadas": 0,
        "confirmadas": 0,
        "no_confirmadas": 0,
        "tasa": None,
        "sin_desenlace": {
            "total": 0,
            "antiguedad_promedio_dias": None,
            "antiguedad_maxima_dias": None,
        },
    }


def test_conversion_cohorte_no_por_fecha_de_confirmacion(client, db, entorno, auth_headers):
    """A mano, con historial sintético: cotizada el 10-mar y confirmada FUERA
    del periodo (abril). Filtrando marzo: cotizadas=1, confirmadas=1 (la
    cohorte sigue a SUS solicitudes) → tasa = 1.0."""
    _sintetica(
        db,
        entorno,
        estado=Estado.CONFIRMADA,
        monto_confirmado=Decimal("100.00"),
        moneda_confirmada=Moneda.MXN,
        confirmado_en=_utc(2026, 4, 2),
        eventos=[
            (Estado.EN_PROCESO, Estado.COTIZADA, _utc(2026, 3, 10)),
            (Estado.COTIZADA, Estado.CONFIRMADA, _utc(2026, 4, 2)),
        ],
    )
    marzo = {"desde": "2026-03-01", "hasta": "2026-03-31"}
    conv = client.get(
        f"{METRICAS}/resumen", params=marzo, headers=auth_headers(entorno.admin)
    ).json()["conversion"]
    assert (conv["cotizadas"], conv["confirmadas"], conv["tasa"]) == (1, 1, 1.0)
    # Correcciones del comprador (eventos de==a) NO inflan el denominador.
    db.add(
        HistorialEstado(
            solicitud_id=_sintetica(
                db,
                entorno,
                estado=Estado.COTIZADA,
                cotizado_en=_utc(2026, 3, 12),
                eventos=[(Estado.EN_PROCESO, Estado.COTIZADA, _utc(2026, 3, 12))],
            ).id,
            de=Estado.COTIZADA,
            a=Estado.COTIZADA,
            usuario_id=entorno.comprador.id,
            timestamp=_utc(2026, 3, 13),
        )
    )
    db.commit()
    conv = client.get(
        f"{METRICAS}/resumen", params=marzo, headers=auth_headers(entorno.admin)
    ).json()["conversion"]
    assert (conv["cotizadas"], conv["confirmadas"]) == (2, 1)
    assert conv["tasa"] == 0.5


# ----------------------------------------------------------- p.2 impresiones


def test_impresion_por_estatus_y_bitacora(client, db, entorno, auth_headers, con_comprobante):
    sid = _enviada(client, entorno, auth_headers)
    headers_v = auth_headers(entorno.vendedor)

    # Antes de COTIZADA no hay documento (el botón vive inactivo en la UI).
    r = client.post(
        f"{BASE}/{sid}/impresiones", headers=headers_v, json={"documento": "COTIZACION"}
    )
    assert r.status_code == 422 and r.json()["code"] == "impresion_no_disponible"

    _capturar_a(client, entorno, auth_headers, sid)
    assert (
        client.post(
            f"{BASE}/{sid}/cotizar", headers=auth_headers(entorno.comprador), json={}
        ).status_code
        == 200
    )
    # COTIZADA → el documento es la Cotización; el Pedido confirmado NO existe.
    r = client.post(
        f"{BASE}/{sid}/impresiones", headers=headers_v, json={"documento": "COTIZACION"}
    )
    assert r.status_code == 201, r.text
    r = client.post(
        f"{BASE}/{sid}/impresiones", headers=headers_v, json={"documento": "PEDIDO_CONFIRMADO"}
    )
    assert r.status_code == 422 and r.json()["code"] == "impresion_no_disponible"

    con_comprobante(sid, entorno.vendedor)
    assert (
        client.post(f"{BASE}/{sid}/seleccionar", headers=headers_v, json={"letra": "A"}).status_code
        == 200
    )
    # CONFIRMADA → Pedido confirmado; la Cotización sigue REIMPRIMIBLE.
    r = client.post(
        f"{BASE}/{sid}/impresiones", headers=headers_v, json={"documento": "PEDIDO_CONFIRMADO"}
    )
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["documento"] == "PEDIDO_CONFIRMADO" and body["estado"] == "CONFIRMADA"
    r = client.post(
        f"{BASE}/{sid}/impresiones",
        headers=auth_headers(entorno.comprador),
        json={"documento": "COTIZACION"},
    )
    assert r.status_code == 201, r.text

    # Bitácora: snapshot autosuficiente (folio, usuario y rol como texto).
    filas = list(db.scalars(select(Impresion).where(Impresion.solicitud_id == sid)))
    assert [f.documento for f in filas] == ["COTIZACION", "PEDIDO_CONFIRMADO", "COTIZACION"]
    assert filas[0].usuario == entorno.vendedor.nombre and filas[0].rol == "vendedor"
    assert filas[2].rol == "comprador"
    assert all(f.folio and f.creado_en is not None for f in filas)


def test_impresion_respeta_scoping(client, entorno, auth_headers, con_comprobante):
    """El ajeno no registra impresiones de solicitudes que no ve (mismo 404
    del detalle: la existencia no se filtra)."""
    sid = _cotizada(client, entorno, auth_headers)
    r = client.post(
        f"{BASE}/{sid}/impresiones",
        headers=auth_headers(entorno.ajeno),
        json={"documento": "COTIZACION"},
    )
    assert r.status_code == 404


# ------------------------------------------------- p.4 export de comparativas


@pytest.fixture
def tres_sucursales(db, entorno, make_sucursal):
    """Dataset a mano para el export: A y B con dinero confirmado (569,939.56
    y 78,928.45 — el orden alfabético las invertiría), C sin dinero (vacío) y
    con un ciclo abierto ROJO (10 días)."""
    suc_a, suc_b, suc_c = (make_sucursal(n) for n in ("F14 A", "F14 B", "F14 C"))
    ahora = datetime.now(UTC)

    def _solicitud(sucursal, estado, eventos, **campos):
        s = Solicitud(
            folio=f"F14-{next(_folio)}",
            vendedor_id=entorno.vendedor.id,
            sucursal_id=sucursal.id,
            comprador_id=entorno.comprador.id,
            estado=estado,
            prioridad=Prioridad.NORMAL,
            **campos,
        )
        db.add(s)
        db.flush()
        for de, a, ts in eventos:
            db.add(
                HistorialEstado(
                    solicitud_id=s.id, de=de, a=a, usuario_id=entorno.vendedor.id, timestamp=ts
                )
            )
        return s

    # A y B: ciclo cerrado de 2 días naturales + confirmadas con monto.
    for sucursal, monto in ((suc_a, "569939.56"), (suc_b, "78928.45")):
        _solicitud(
            sucursal,
            Estado.CONFIRMADA,
            [
                (Estado.BORRADOR, Estado.ENVIADA, ahora - timedelta(days=3)),
                (Estado.EN_PROCESO, Estado.COTIZADA, ahora - timedelta(days=2)),
            ],
            monto_confirmado=Decimal(monto),
            moneda_confirmada=Moneda.MXN,
            confirmado_en=ahora - timedelta(days=1),
        )
    # C: ciclo ABIERTO desde hace 10 días → banda LENTA (roja).
    _solicitud(
        suc_c,
        Estado.ENVIADA,
        [(Estado.BORRADOR, Estado.ENVIADA, ahora - timedelta(days=10))],
    )
    db.commit()
    return SimpleNamespace(a=suc_a, b=suc_b, c=suc_c)


def _filas_export(ws):
    encabezados = [c.value for c in ws[6]]
    filas = list(ws.iter_rows(min_row=7, values_only=True))
    return encabezados, filas


def test_export_comparativas_tipos_orden_y_encabezado(
    client, entorno, tres_sucursales, auth_headers
):
    r = client.get(
        EXPORT_COMP,
        params={"dimension": "sucursal", "orden": "confirmado_mxn", "direccion": "desc", **HOY},
        headers=auth_headers(entorno.admin),
    )
    assert r.status_code == 200, r.text
    assert "Comparativas_Por-sucursal_2026-01_" in r.headers["content-disposition"]
    ws = load_workbook(BytesIO(r.content)).active

    # Encabezado del reporte: nombre, pestaña, periodo y generación.
    assert ws["A1"].value == "Comparativas — Sistema de Cotizaciones Herinox"
    assert (ws["A2"].value, ws["B2"].value) == ("Pestaña", "Por sucursal")
    assert ws["B3"].value == "2026-01-01 a 2030-12-31"
    assert "UTC" in ws["B4"].value

    encabezados, filas = _filas_export(ws)
    assert encabezados[:3] == ["Nombre", "Volumen", "Ciclos cerrados"]
    col = {t: i for i, t in enumerate(encabezados)}
    nuestras = [f for f in filas if str(f[0]).startswith("F14 ")]

    # ORDEN numérico desc por Confirmado (MXN): 569,939.56 > 78,928.45 (el
    # orden alfabético diría lo contrario); el vacío (C) SIEMPRE al final.
    assert [f[0] for f in nuestras] == ["F14 A", "F14 B", "F14 C"]
    # TIPOS reales de celda: dinero numérico con formato de moneda, % como
    # porcentaje, enteros como número — nada como texto.
    fila_a = nuestras[0]
    assert float(fila_a[col["Confirmado (MXN)"]]) == 569939.56
    assert fila_a[col["Volumen"]] == 1 and isinstance(fila_a[col["Volumen"]], int)
    assert float(fila_a[col["% banda esperada"]]) == 1.0  # 1/1 ciclos en esperada... a mano
    celda_dinero = ws.cell(row=7, column=col["Confirmado (MXN)"] + 1)
    assert celda_dinero.number_format == "$#,##0.00"
    celda_pct = ws.cell(row=7, column=col["% banda esperada"] + 1)
    assert celda_pct.number_format == "0.0%"
    assert fila_a[col["Confirmado (MXN)"]] is not None
    # El vacío es celda VACÍA (None), no el texto "—".
    assert nuestras[2][col["Confirmado (MXN)"]] is None

    # Dirección asc: B, A y el vacío (C) SIGUE al final.
    r = client.get(
        EXPORT_COMP,
        params={"dimension": "sucursal", "orden": "confirmado_mxn", "direccion": "asc", **HOY},
        headers=auth_headers(entorno.admin),
    )
    _, filas = _filas_export(load_workbook(BytesIO(r.content)).active)
    assert [f[0] for f in filas if str(f[0]).startswith("F14 ")] == ["F14 B", "F14 A", "F14 C"]


def test_export_comparativas_orden_rojas(client, entorno, tres_sucursales, auth_headers):
    """Criterio V/A/R: rojas DESC pone lo urgente arriba (C, con su ciclo
    abierto de 10 días, encabeza)."""
    r = client.get(
        EXPORT_COMP,
        params={"dimension": "sucursal", "orden": "rojas", "direccion": "desc", **HOY},
        headers=auth_headers(entorno.admin),
    )
    encabezados, filas = _filas_export(load_workbook(BytesIO(r.content)).active)
    nuestras = [f for f in filas if str(f[0]).startswith("F14 ")]
    col = {t: i for i, t in enumerate(encabezados)}
    assert nuestras[0][0] == "F14 C" and nuestras[0][col["Rojas"]] == 1
    assert [f[col["Rojas"]] for f in nuestras] == [1, 0, 0]


def test_export_comparativas_gates_y_dinero_por_rol(client, entorno, tres_sucursales, auth_headers):
    # Gates por dimensión = los de las tablas.
    r = client.get(
        EXPORT_COMP,
        params={"dimension": "vendedor"},
        headers=auth_headers(entorno.vendedor),
    )
    assert r.status_code == 403
    r = client.get(
        EXPORT_COMP,
        params={"dimension": "comprador"},
        headers=auth_headers(entorno.comprador),
    )
    assert r.status_code == 403
    r = client.get(
        EXPORT_COMP,
        params={"dimension": "sucursal", "orden": "nada"},
        headers=auth_headers(entorno.admin),
    )
    assert r.status_code == 422 and r.json()["code"] == "orden_invalido"

    # §0 heredado: el VENDEDOR exporta por-sucursal SIN columna de dinero.
    r = client.get(
        EXPORT_COMP,
        params={"dimension": "sucursal", **HOY},
        headers=auth_headers(entorno.vendedor),
    )
    assert r.status_code == 200, r.text
    encabezados, _ = _filas_export(load_workbook(BytesIO(r.content)).active)
    assert "Confirmado (MXN)" not in encabezados
    # El gerente de sucursal SÍ la lleva (§4.9: gerentes ven consolidados).
    r = client.get(
        EXPORT_COMP,
        params={"dimension": "vendedor", **HOY},
        headers=auth_headers(entorno.gerente),
    )
    encabezados, _ = _filas_export(load_workbook(BytesIO(r.content)).active)
    assert "Confirmado (MXN)" in encabezados
