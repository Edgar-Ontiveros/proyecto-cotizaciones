"""F8f: tiempos por etapa (segmentos desde el historial) y solicitudes de
PROYECTO.

Tiempos: historial SINTÉTICO de timestamps controlados y aritmética A MANO en
los docstrings. Sucursal base America/Chihuahua (UTC-6 fijo, sin DST desde
2022); jornada L–V 08:00–18:00 y sábado 08:00–13:00. Calendario de marzo
2026: el 1 es domingo; 02=lun, 03=mar, 05=jue. Sin festivos en las fechas
usadas.
"""

from datetime import UTC, datetime
from types import SimpleNamespace

import pytest

from app.models.historial import HistorialEstado
from app.models.notificacion import Notificacion
from app.models.solicitud import Estado, Prioridad, Solicitud
from app.models.sucursal import CompradorSucursal
from app.models.usuario import Rol
from app.modules.metricas.tiempos import cargar_tiempos

BASE = "/api/v1/solicitudes"
TIEMPOS_ETAPA = "/api/v1/metricas/tiempos-etapa"
PARTIDA = {"cantidad": "2", "unidad": "PZ", "descripcion": "PTR 2X2"}
MARZO = {"desde": "2026-03-01", "hasta": "2026-03-31"}


@pytest.fixture
def entorno(db, make_user, make_sucursal):
    cuu = make_sucursal("F8f CUU")  # America/Chihuahua (default del factory)
    tij = make_sucursal("F8f TIJ")
    tij.timezone = "America/Tijuana"
    comprador = make_user(Rol.COMPRADOR)
    db.add(CompradorSucursal(comprador_id=comprador.id, sucursal_id=cuu.id, titular=True))
    db.commit()
    return SimpleNamespace(
        cuu=cuu,
        tij=tij,
        comprador=comprador,
        vendedor=make_user(Rol.VENDEDOR, sucursal_id=cuu.id),
        admin=make_user(Rol.ADMIN),
    )


_folio = iter(range(1, 10_000))


def _sintetica(db, entorno, *, estado, eventos, sucursal=None, vendedor=None):
    """Solicitud con historial sintético; creado_en = timestamp del primer
    evento (en producción el nacimiento comparte transacción con creado_en)."""
    solicitud = Solicitud(
        folio=f"F8F-{next(_folio)}",
        vendedor_id=(vendedor or entorno.vendedor).id,
        sucursal_id=(sucursal or entorno.cuu).id,
        comprador_id=entorno.comprador.id,
        estado=estado,
        prioridad=Prioridad.NORMAL,
    )
    db.add(solicitud)
    db.flush()
    solicitud.creado_en = eventos[0][2]
    for de, a, ts in eventos:
        db.add(
            HistorialEstado(
                solicitud_id=solicitud.id,
                de=de,
                a=a,
                usuario_id=solicitud.vendedor_id,
                timestamp=ts,
            )
        )
    db.commit()
    return solicitud


def _utc(d, hh, mm=0):
    return datetime(2026, 3, d, hh, mm, tzinfo=UTC)


# ------------------------------------------------------------------- tiempos


def _flujo_completo(db, entorno):
    """Flujo multi-estado, todo en horario hábil (local = UTC-6):

    lun 02: 09:00 nace BORRADOR · 11:00 ENVIADA · 15:00 EN_PROCESO
    mar 03: 10:00 COTIZADA · 14:00 CONFIRMADA

    Segmentos (háb | nat):
    - BORRADOR   lun 09→11            = 2  | 2
    - ENVIADA    lun 11→15            = 4  | 4
    - EN_PROCESO lun 15→18 + mar 8→10 = 5  | 19 (15:00 lun → 10:00 mar)
    - COTIZADA   mar 10→14            = 4  | 4
    - CONFIRMADA vigente (terminal)   = 0  | 0  (temporizador detenido)

    General = 2+4+5+4 = 15 háb / 2+4+19+4 = 29 nat.
    Compras (ENVIADA+EN_PROCESO) = 4+5 = 9. Ventas (BORRADOR+COTIZADA) = 2+4 = 6.
    """
    return _sintetica(
        db,
        entorno,
        estado=Estado.CONFIRMADA,
        eventos=[
            (None, Estado.BORRADOR, _utc(2, 15)),
            (Estado.BORRADOR, Estado.ENVIADA, _utc(2, 17)),
            (Estado.ENVIADA, Estado.EN_PROCESO, _utc(2, 21)),
            (Estado.EN_PROCESO, Estado.COTIZADA, _utc(3, 16)),
            (Estado.COTIZADA, Estado.CONFIRMADA, _utc(3, 20)),
        ],
    )


def test_flujo_completo_multi_estado(db, entorno):
    """Aritmética del docstring de _flujo_completo, verificada segmento a
    segmento y en los tres agregados."""
    solicitud = _flujo_completo(db, entorno)
    t = cargar_tiempos(db, [solicitud.id], ahora=_utc(5, 12)).get(solicitud.id)
    assert t is not None
    assert [(s.estado, s.horas_habiles, s.horas_naturales) for s in t.segmentos] == [
        (Estado.BORRADOR, 2.0, 2.0),
        (Estado.ENVIADA, 4.0, 4.0),
        (Estado.EN_PROCESO, 5.0, 19.0),
        (Estado.COTIZADA, 4.0, 4.0),
        (Estado.CONFIRMADA, 0.0, 0.0),
    ]
    assert t.segmentos[-1].fin is None  # vigente, pero detenido
    assert t.general_horas_habiles == 15.0
    assert t.general_horas_naturales == 29.0
    assert t.compras_horas_habiles == 9.0
    assert t.ventas_horas_habiles == 6.0
    assert t.detenido is True


def test_reenvio_suma_rechazada_a_ventas_y_dos_enviadas_a_compras(db, entorno):
    """Reenvío, todo el lunes 02 (local = UTC-6):

    08:00 nace · 10:00 ENVIADA · 12:00 RECHAZADA · 14:00 ENVIADA (reenvío)
    · 16:00 EN_PROCESO · 17:00 COTIZADA · ahora = 19:00 (fuera de jornada).

    - BORRADOR 8→10 = 2 · ENVIADA 10→12 = 2 · RECHAZADA 12→14 = 2
    - ENVIADA 14→16 = 2 · EN_PROCESO 16→17 = 1
    - COTIZADA vigente 17→ahora: háb 17→18 = 1, nat 2.
    Compras = 2+2+1 = 5 (DOS pasadas por ENVIADA suman).
    Ventas = 2 (BORRADOR) + 2 (RECHAZADA) + 1 (COTIZADA) = 5.
    General = 10 háb / 11 nat. Sigue viva: detenido=False.
    """
    solicitud = _sintetica(
        db,
        entorno,
        estado=Estado.COTIZADA,
        eventos=[
            (None, Estado.BORRADOR, _utc(2, 14)),
            (Estado.BORRADOR, Estado.ENVIADA, _utc(2, 16)),
            (Estado.ENVIADA, Estado.RECHAZADA, _utc(2, 18)),
            (Estado.RECHAZADA, Estado.ENVIADA, _utc(2, 20)),
            (Estado.ENVIADA, Estado.EN_PROCESO, _utc(2, 22)),
            (Estado.EN_PROCESO, Estado.COTIZADA, _utc(2, 23)),
        ],
    )
    t = cargar_tiempos(db, [solicitud.id], ahora=_utc(3, 1)).get(solicitud.id)
    assert t is not None
    assert t.compras_horas_habiles == 5.0
    assert t.ventas_horas_habiles == 5.0
    assert t.general_horas_habiles == 10.0
    assert t.general_horas_naturales == 11.0
    assert t.detenido is False
    enviadas = [s for s in t.segmentos if s.estado == Estado.ENVIADA]
    assert len(enviadas) == 2 and all(s.horas_habiles == 2.0 for s in enviadas)


def test_eventos_de_igual_a_no_cortan_segmentos(db, entorno):
    """Lunes 02: nace 08:00 · ENVIADA 10:00 · edición (de==a) 11:00 ·
    RECHAZADA 12:00. La edición NO corta: 3 segmentos, ENVIADA 10→12 = 2 háb."""
    solicitud = _sintetica(
        db,
        entorno,
        estado=Estado.RECHAZADA,
        eventos=[
            (None, Estado.BORRADOR, _utc(2, 14)),
            (Estado.BORRADOR, Estado.ENVIADA, _utc(2, 16)),
            (Estado.ENVIADA, Estado.ENVIADA, _utc(2, 17)),  # edición
            (Estado.ENVIADA, Estado.RECHAZADA, _utc(2, 18)),
        ],
    )
    t = cargar_tiempos(db, [solicitud.id], ahora=_utc(2, 19)).get(solicitud.id)
    assert t is not None
    assert [s.estado for s in t.segmentos] == [
        Estado.BORRADOR,
        Estado.ENVIADA,
        Estado.RECHAZADA,
    ]
    assert t.segmentos[1].horas_habiles == 2.0


def test_terminal_detiene_el_temporizador_general(db, entorno):
    """Lunes 02: nace 08:00 · ENVIADA 10:00 · CANCELADA 12:00; ahora = jueves.
    General queda CONGELADO en 2+2 = 4 háb / 4 nat aunque pasen días; el
    segmento CANCELADA vigente no mide nada."""
    solicitud = _sintetica(
        db,
        entorno,
        estado=Estado.CANCELADA,
        eventos=[
            (None, Estado.BORRADOR, _utc(2, 14)),
            (Estado.BORRADOR, Estado.ENVIADA, _utc(2, 16)),
            (Estado.ENVIADA, Estado.CANCELADA, _utc(2, 18)),
        ],
    )
    t = cargar_tiempos(db, [solicitud.id], ahora=_utc(5, 20)).get(solicitud.id)
    assert t is not None
    assert t.general_horas_habiles == 4.0
    assert t.general_horas_naturales == 4.0
    assert t.detenido is True
    assert t.segmentos[-1].estado == Estado.CANCELADA
    assert t.segmentos[-1].horas_habiles == 0.0


def test_reversion_de_no_confirmada_reanuda_el_general(db, entorno):
    """BORDE documentado: la pausa en NO_CONFIRMADA NO cuenta y la reversión
    reanuda el reloj.

    lun 02 (local): 08:00 nace · 09:00 ENVIADA · 10:00 EN_PROCESO · 11:00
    COTIZADA · 13:00 NO_CONFIRMADA. mar 03 10:00: admin revierte a COTIZADA.
    ahora = mar 12:00.

    - BORRADOR 1 · ENVIADA 1 · EN_PROCESO 1 · COTIZADA 11→13 = 2
    - NO_CONFIRMADA (CERRADA) lun 13→18 = 5 + mar 8→10 = 2 → 7 háb
      informativas, EXCLUIDAS de todo agregado.
    - COTIZADA vigente mar 10→12 = 2.
    General = 1+1+1+2+2 = 7 háb (sin las 7 de la pausa). Ventas = 1+2+2 = 5.
    Compras = 1+1 = 2. detenido=False (reanudado).
    """
    solicitud = _sintetica(
        db,
        entorno,
        estado=Estado.COTIZADA,
        eventos=[
            (None, Estado.BORRADOR, _utc(2, 14)),
            (Estado.BORRADOR, Estado.ENVIADA, _utc(2, 15)),
            (Estado.ENVIADA, Estado.EN_PROCESO, _utc(2, 16)),
            (Estado.EN_PROCESO, Estado.COTIZADA, _utc(2, 17)),
            (Estado.COTIZADA, Estado.NO_CONFIRMADA, _utc(2, 19)),
            (Estado.NO_CONFIRMADA, Estado.COTIZADA, _utc(3, 16)),
        ],
    )
    t = cargar_tiempos(db, [solicitud.id], ahora=_utc(3, 18)).get(solicitud.id)
    assert t is not None
    pausa = next(s for s in t.segmentos if s.estado == Estado.NO_CONFIRMADA)
    assert pausa.fin is not None and pausa.horas_habiles == 7.0
    assert t.general_horas_habiles == 7.0
    assert t.ventas_horas_habiles == 5.0
    assert t.compras_horas_habiles == 2.0
    assert t.detenido is False


def test_multi_tz_mismos_instantes_distintas_horas(db, entorno):
    """Mismos instantes UTC, sucursales distintas. Lunes 02, 15:00Z:
    CUU (UTC-6) son las 09:00 — ya en jornada; TIJ (UTC-8, sin DST hasta el
    08-mar) son las 07:00 — antes de jornada.

    nace 15:00Z · ENVIADA 17:00Z · ahora 19:00Z.
    BORRADOR: CUU 09→11 = 2 háb; TIJ 08→09 = 1 háb (07:00 no cuenta).
    ENVIADA vigente: CUU 11→13 = 2; TIJ 09→11 = 2.
    """
    eventos = [
        (None, Estado.BORRADOR, _utc(2, 15)),
        (Estado.BORRADOR, Estado.ENVIADA, _utc(2, 17)),
    ]
    en_cuu = _sintetica(db, entorno, estado=Estado.ENVIADA, eventos=eventos)
    en_tij = _sintetica(db, entorno, estado=Estado.ENVIADA, eventos=eventos, sucursal=entorno.tij)
    tiempos = cargar_tiempos(db, [en_cuu.id, en_tij.id], ahora=_utc(2, 19))
    assert tiempos[en_cuu.id].ventas_horas_habiles == 2.0
    assert tiempos[en_tij.id].ventas_horas_habiles == 1.0
    assert tiempos[en_cuu.id].compras_horas_habiles == 2.0
    assert tiempos[en_tij.id].compras_horas_habiles == 2.0
    assert tiempos[en_cuu.id].general_horas_naturales == 4.0
    assert tiempos[en_tij.id].general_horas_naturales == 4.0


def test_detalle_incluye_bloque_tiempos_para_todo_rol(client, db, entorno, auth_headers):
    """El bloque `tiempos` va en el detalle para TODO rol con acceso (no hay
    dinero). Flujo completo: general 15 háb / 29 nat, compras 9, ventas 6."""
    solicitud = _flujo_completo(db, entorno)
    for usuario in (entorno.vendedor, entorno.comprador, entorno.admin):
        r = client.get(f"{BASE}/{solicitud.id}", headers=auth_headers(usuario))
        assert r.status_code == 200, r.text
        tiempos = r.json()["tiempos"]
        assert tiempos["general_horas_habiles"] == 15.0
        assert tiempos["general_horas_naturales"] == 29.0
        assert tiempos["compras_horas_habiles"] == 9.0
        assert tiempos["ventas_horas_habiles"] == 6.0
        assert tiempos["detenido"] is True
        assert [s["estado"] for s in tiempos["segmentos"]] == [
            "BORRADOR",
            "ENVIADA",
            "EN_PROCESO",
            "COTIZADA",
            "CONFIRMADA",
        ]


def test_export_columnas_de_tiempos(client, db, entorno, auth_headers):
    """Export del flujo completo: total general natural 29, compras háb 9,
    ventas háb 6 (columnas nuevas tras 'Horas hábiles último ciclo')."""
    from io import BytesIO

    from openpyxl import load_workbook

    solicitud = _flujo_completo(db, entorno)
    r = client.get("/api/v1/solicitudes/export", headers=auth_headers(entorno.admin))
    assert r.status_code == 200
    ws = load_workbook(BytesIO(r.content)).active
    encabezados = [c.value for c in ws[1]]
    col = {nombre: i for i, nombre in enumerate(encabezados)}
    fila = next(f for f in ws.iter_rows(min_row=2, values_only=True) if f[0] == solicitud.folio)
    assert fila[col["Total general (hrs naturales)"]] == 29.0
    assert fila[col["Tiempo compras (hrs hábiles)"]] == 9.0
    assert fila[col["Tiempo ventas (hrs hábiles)"]] == 6.0


def test_metricas_tiempos_etapa(client, db, entorno, auth_headers):
    """Solo los segmentos CERRADOS alimentan promedio/mediana.

    A = flujo completo: BORRADOR 2, ENVIADA 4, EN_PROCESO 5, COTIZADA 4
    (cerrados; compras 9, ventas 6).
    B = nace lun 08:00, ENVIADA 09:00 (vigente): BORRADOR cerrado 1; su
    ENVIADA abierta NO cuenta.

    ENVIADA: n=1 (solo A) → prom=med=4. BORRADOR: n=2, obs {2,1} → prom=med=1.5.
    Compras: n=1 (solo A) → 9. Ventas: n=2, obs {6,1} → prom=med=3.5.
    """
    _flujo_completo(db, entorno)
    _sintetica(
        db,
        entorno,
        estado=Estado.ENVIADA,
        eventos=[
            (None, Estado.BORRADOR, _utc(2, 14)),
            (Estado.BORRADOR, Estado.ENVIADA, _utc(2, 15)),
        ],
    )
    r = client.get(TIEMPOS_ETAPA, params=MARZO, headers=auth_headers(entorno.admin))
    assert r.status_code == 200, r.text
    datos = r.json()
    assert datos["por_estado"]["ENVIADA"] == {
        "n": 1,
        "promedio_horas_habiles": 4.0,
        "mediana_horas_habiles": 4.0,
    }
    assert datos["por_estado"]["BORRADOR"] == {
        "n": 2,
        "promedio_horas_habiles": 1.5,
        "mediana_horas_habiles": 1.5,
    }
    assert datos["por_estado"]["EN_PROCESO"]["n"] == 1
    assert datos["compras"] == {
        "n": 1,
        "promedio_horas_habiles": 9.0,
        "mediana_horas_habiles": 9.0,
    }
    assert datos["ventas"] == {
        "n": 2,
        "promedio_horas_habiles": 3.5,
        "mediana_horas_habiles": 3.5,
    }


def test_tiempos_etapa_scoping_vendedor(client, db, entorno, auth_headers, make_user):
    """Scoping estándar de /metricas: el vendedor solo ve lo suyo."""
    _flujo_completo(db, entorno)  # del vendedor del entorno
    otro = make_user(Rol.VENDEDOR, sucursal_id=entorno.cuu.id)
    _sintetica(
        db,
        entorno,
        estado=Estado.RECHAZADA,
        vendedor=otro,
        eventos=[
            (None, Estado.BORRADOR, _utc(2, 14)),
            (Estado.BORRADOR, Estado.ENVIADA, _utc(2, 15)),
            (Estado.ENVIADA, Estado.RECHAZADA, _utc(2, 16)),
        ],
    )
    r = client.get(TIEMPOS_ETAPA, params=MARZO, headers=auth_headers(otro))
    assert r.status_code == 200
    datos = r.json()
    # Solo SU solicitud: una pasada cerrada por ENVIADA de 1h.
    assert datos["por_estado"]["ENVIADA"] == {
        "n": 1,
        "promedio_horas_habiles": 1.0,
        "mediana_horas_habiles": 1.0,
    }
    assert datos["por_estado"]["EN_PROCESO"]["n"] == 0


# ------------------------------------------------------------------ proyecto


def _crear(client, headers, **extra):
    r = client.post(
        BASE, headers=headers, json={"cliente": "DINCO", "partidas": [PARTIDA], **extra}
    )
    assert r.status_code == 201, r.text
    return r.json()


def _notifs(db, solicitud_id):
    from sqlalchemy import select

    filas = db.scalars(select(Notificacion).where(Notificacion.solicitud_id == solicitud_id)).all()
    return [(n.usuario_id, n.tipo, n.mensaje) for n in filas]


def test_es_proyecto_se_define_al_crear_y_solo_cambia_en_borrador(client, entorno, auth_headers):
    headers = auth_headers(entorno.vendedor)
    creada = _crear(client, headers, es_proyecto=True)
    assert creada["es_proyecto"] is True

    # En BORRADOR sí cambia (ida y vuelta).
    r = client.patch(
        f"{BASE}/{creada['id']}",
        headers=headers,
        json={"cliente": "DINCO", "partidas": [PARTIDA], "es_proyecto": False},
    )
    assert r.status_code == 200 and r.json()["es_proyecto"] is False
    r = client.patch(
        f"{BASE}/{creada['id']}",
        headers=headers,
        json={"cliente": "DINCO", "partidas": [PARTIDA], "es_proyecto": True},
    )
    assert r.status_code == 200 and r.json()["es_proyecto"] is True

    assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200
    # Fuera de BORRADOR: cambiarlo → 422; omitirlo o repetirlo → sin cambio.
    r = client.patch(
        f"{BASE}/{creada['id']}",
        headers=headers,
        json={"cliente": "DINCO", "partidas": [PARTIDA], "es_proyecto": False},
    )
    assert r.status_code == 422 and r.json()["code"] == "es_proyecto_inmutable"
    r = client.patch(
        f"{BASE}/{creada['id']}",
        headers=headers,
        json={"cliente": "DINCO", "partidas": [PARTIDA]},
    )
    assert r.status_code == 200 and r.json()["es_proyecto"] is True
    r = client.patch(
        f"{BASE}/{creada['id']}",
        headers=headers,
        json={"cliente": "DINCO", "partidas": [PARTIDA], "es_proyecto": True},
    )
    assert r.status_code == 200 and r.json()["es_proyecto"] is True


def test_notificaciones_exactas_al_enviar_proyecto(client, db, entorno, auth_headers, make_user):
    gc1 = make_user(Rol.GERENTE_COMPRAS)
    gc2 = make_user(Rol.GERENTE_COMPRAS)
    gc_inactivo = make_user(Rol.GERENTE_COMPRAS, activo=False)
    gerente = make_user(Rol.GERENTE_SUCURSAL, sucursal_id=entorno.cuu.id)
    gerente_otra = make_user(Rol.GERENTE_SUCURSAL, sucursal_id=entorno.tij.id)

    headers = auth_headers(entorno.vendedor)
    creada = _crear(client, headers, es_proyecto=True)
    assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200

    notifs = _notifs(db, creada["id"])
    folio = client.get(f"{BASE}/{creada['id']}", headers=headers).json()["folio"]
    esperadas = {
        (entorno.comprador.id, "asignacion", f"Se te asignó la solicitud {folio}"),
        (gc1.id, "proyecto_compras", f"Nueva solicitud de PROYECTO {folio}"),
        (gc2.id, "proyecto_compras", f"Nueva solicitud de PROYECTO {folio}"),
        (gerente.id, "proyecto_sucursal", f"Nueva solicitud de PROYECTO {folio}"),
    }
    # Exactas y NADIE más: ni el gerente de otra sucursal, ni el inactivo,
    # ni admin.
    assert set(notifs) == esperadas and len(notifs) == 4
    assert gc_inactivo.id not in {u for u, _, _ in notifs}
    assert gerente_otra.id not in {u for u, _, _ in notifs}


def test_proyecto_sucursal_sin_gerente_solo_notifica_compras(
    client, db, entorno, auth_headers, make_user
):
    """TIK/Manufactura no tienen gerente: solo la notificación normal + los
    gerentes de compras."""
    gc = make_user(Rol.GERENTE_COMPRAS)
    headers = auth_headers(entorno.vendedor)
    creada = _crear(client, headers, es_proyecto=True)
    assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200
    tipos = sorted(t for _, t, _ in _notifs(db, creada["id"]))
    assert tipos == ["asignacion", "proyecto_compras"]
    assert (gc.id, "proyecto_compras") in {(u, t) for u, t, _ in _notifs(db, creada["id"])}


def test_reenvio_de_proyecto_re_notifica(client, db, entorno, auth_headers, make_user):
    gc = make_user(Rol.GERENTE_COMPRAS)
    gerente = make_user(Rol.GERENTE_SUCURSAL, sucursal_id=entorno.cuu.id)
    headers = auth_headers(entorno.vendedor)
    creada = _crear(client, headers, es_proyecto=True)
    assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200

    # Rechazo del comprador y reenvío del vendedor.
    from app.models.catalogos import FamiliaMotivo, MotivoRechazo

    motivo = MotivoRechazo(familia=FamiliaMotivo.FALTA_INFORMACION, texto="Motivo F8f")
    db.add(motivo)
    db.commit()
    r = client.post(
        f"{BASE}/{creada['id']}/rechazar",
        headers=auth_headers(entorno.comprador),
        json={"motivo_id": motivo.id},
    )
    assert r.status_code == 200, r.text
    assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200

    notifs = _notifs(db, creada["id"])
    reenvios = [m for _, t, m in notifs if t == "proyecto_compras" and "reenviada" in m]
    assert len(reenvios) == 1  # el reenvío re-notificó a compras...
    assert sum(1 for _, t, _ in notifs if t == "proyecto_compras") == 2
    assert sum(1 for u, t, _ in notifs if t == "proyecto_sucursal" and u == gerente.id) == 2
    assert gc.id in {u for u, t, _ in notifs if t == "proyecto_compras"}


def test_gerente_no_se_auto_notifica_de_su_propio_proyecto(
    client, db, entorno, auth_headers, make_user
):
    """v3: el gerente crea y envía; su propia notificación especial sobraría."""
    gerente = make_user(Rol.GERENTE_SUCURSAL, sucursal_id=entorno.cuu.id)
    headers = auth_headers(gerente)
    creada = _crear(client, headers, es_proyecto=True)
    assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200
    assert "proyecto_sucursal" not in {t for _, t, _ in _notifs(db, creada["id"])}


def test_solicitud_normal_no_genera_notificaciones_de_proyecto(
    client, db, entorno, auth_headers, make_user
):
    make_user(Rol.GERENTE_COMPRAS)
    make_user(Rol.GERENTE_SUCURSAL, sucursal_id=entorno.cuu.id)
    headers = auth_headers(entorno.vendedor)
    creada = _crear(client, headers)  # sin es_proyecto → False
    assert creada["es_proyecto"] is False
    assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200
    tipos = {t for _, t, _ in _notifs(db, creada["id"])}
    assert tipos == {"asignacion"}


def test_historial_del_envio_menciona_proyecto(client, entorno, auth_headers):
    headers = auth_headers(entorno.vendedor)
    creada = _crear(client, headers, es_proyecto=True)
    assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200
    detalle = client.get(f"{BASE}/{creada['id']}", headers=headers).json()
    envio = next(e for e in detalle["historial"] if e["a"] == "ENVIADA")
    assert envio["comentario"] == "Solicitud de PROYECTO"


def test_filtro_es_proyecto_en_listado_y_export(client, entorno, auth_headers):
    from io import BytesIO

    from openpyxl import load_workbook

    headers = auth_headers(entorno.vendedor)
    proyecto = _crear(client, headers, es_proyecto=True)
    normal = _crear(client, headers)
    for creada in (proyecto, normal):
        assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200

    solo_proyectos = client.get(BASE, params={"es_proyecto": True}, headers=headers).json()
    assert [i["id"] for i in solo_proyectos["items"]] == [proyecto["id"]]
    solo_normales = client.get(BASE, params={"es_proyecto": False}, headers=headers).json()
    assert [i["id"] for i in solo_normales["items"]] == [normal["id"]]
    # es_proyecto presente en el schema de TODOS los items del listado.
    todos = client.get(BASE, headers=headers).json()
    assert {i["id"]: i["es_proyecto"] for i in todos["items"]} == {
        proyecto["id"]: True,
        normal["id"]: False,
    }

    r = client.get(
        "/api/v1/solicitudes/export",
        params={"es_proyecto": True},
        headers=auth_headers(entorno.admin),
    )
    assert r.status_code == 200
    ws = load_workbook(BytesIO(r.content)).active
    col = {c.value: i for i, c in enumerate(ws[1])}
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(filas) == 1 and filas[0][col["Proyecto"]] == "Sí"


def test_es_proyecto_en_schemas_de_todos_los_roles(client, db, entorno, auth_headers):
    headers = auth_headers(entorno.vendedor)
    creada = _crear(client, headers, es_proyecto=True)
    assert client.post(f"{BASE}/{creada['id']}/enviar", headers=headers).status_code == 200
    # Vendedor (schema base), comprador (consolidado) y admin: la clave existe.
    for usuario in (entorno.vendedor, entorno.comprador, entorno.admin):
        listado = client.get(BASE, headers=auth_headers(usuario)).json()
        assert all("es_proyecto" in item for item in listado["items"])
        detalle = client.get(f"{BASE}/{creada['id']}", headers=auth_headers(usuario)).json()
        assert detalle["es_proyecto"] is True
