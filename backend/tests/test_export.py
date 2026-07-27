"""F6: listado con ciclo vigente (sin N+1) y export a Excel."""

from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook
from sqlalchemy import event

from app.core.database import engine
from app.models.cliente import Cliente
from app.models.cotizacion import Moneda
from app.models.historial import HistorialEstado
from app.models.solicitud import Estado, Prioridad, Solicitud
from app.models.sucursal import CompradorSucursal
from app.models.usuario import Rol

BASE = "/api/v1/solicitudes"
EXPORT = "/api/v1/solicitudes/export"

PARTIDA = {"cantidad": "2", "unidad": "PZA", "descripcion": "PTR 2X2"}


@pytest.fixture
def entorno(db, make_user, make_sucursal):
    sucursal = make_sucursal("Export Suc")
    comprador = make_user(Rol.COMPRADOR)
    db.add(CompradorSucursal(comprador_id=comprador.id, sucursal_id=sucursal.id, titular=True))
    cliente = Cliente(nombre_normalizado="DINCO")
    db.add(cliente)
    db.commit()
    return SimpleNamespace(
        sucursal=sucursal,
        comprador=comprador,
        cliente=cliente,
        vendedor=make_user(Rol.VENDEDOR, sucursal_id=sucursal.id),
        admin=make_user(Rol.ADMIN),
    )


def _enviada(client, entorno, auth_headers):
    headers = auth_headers(entorno.vendedor)
    r = client.post(BASE, headers=headers, json={"cliente": "DINCO", "partidas": [PARTIDA]})
    sid = r.json()["id"]
    assert client.post(f"{BASE}/{sid}/enviar", headers=headers).status_code == 200
    return sid


# ------------------------------------------------------------------- listado


def test_listado_banda_solo_en_abiertas(client, entorno, auth_headers):
    headers = auth_headers(entorno.vendedor)
    abierta = _enviada(client, entorno, auth_headers)
    r = client.post(BASE, headers=headers, json={"cliente": "DINCO", "partidas": [PARTIDA]})
    borrador = r.json()["id"]

    items = {i["id"]: i for i in client.get(BASE, headers=headers).json()["items"]}
    assert items[abierta]["banda"] == "ESPERADA"  # recién enviada: T=0
    assert items[abierta]["dias_transcurridos"] == 0
    assert items[abierta]["horas_habiles"] is not None
    assert items[borrador]["banda"] is None
    assert items[borrador]["dias_transcurridos"] is None
    assert items[borrador]["horas_habiles"] is None


def test_listado_sin_n_mas_uno(client, entorno, auth_headers):
    """El número de queries del listado NO crece con el número de filas
    abiertas (eventos de apertura en un query para toda la página)."""
    headers = auth_headers(entorno.vendedor)
    _enviada(client, entorno, auth_headers)

    def contar() -> int:
        queries: list[str] = []

        def registrar(conn, cursor, statement, parameters, context, executemany):
            queries.append(statement)

        event.listen(engine, "before_cursor_execute", registrar)
        try:
            assert client.get(BASE, headers=headers).status_code == 200
        finally:
            event.remove(engine, "before_cursor_execute", registrar)
        return len(queries)

    con_una = contar()
    for _ in range(4):
        _enviada(client, entorno, auth_headers)
    con_cinco = contar()
    assert con_cinco == con_una  # queries fijos, sin N+1


# -------------------------------------------------------------------- export


def _hoja(respuesta):
    assert respuesta.status_code == 200, respuesta.text
    assert "attachment" in respuesta.headers["content-disposition"]
    assert "solicitudes_" in respuesta.headers["content-disposition"]
    return load_workbook(BytesIO(respuesta.content)).active


def test_export_valido_con_fechas_locales(client, db, entorno, auth_headers):
    creado = datetime(2026, 3, 5, 21, 0, tzinfo=UTC)  # 15:00 en Chihuahua
    solicitud = Solicitud(
        folio="EXP-1",
        vendedor_id=entorno.vendedor.id,
        comprador_id=entorno.comprador.id,
        sucursal_id=entorno.sucursal.id,
        cliente_id=entorno.cliente.id,
        estado=Estado.CONFIRMADA,
        prioridad=Prioridad.NORMAL,
        monto_confirmado=Decimal("1234.50"),
        moneda_confirmada=Moneda.MXN,
        confirmado_en=creado + timedelta(days=1),
    )
    db.add(solicitud)
    db.flush()
    solicitud.creado_en = creado
    db.add(
        HistorialEstado(
            solicitud_id=solicitud.id,
            de=Estado.BORRADOR,
            a=Estado.ENVIADA,
            usuario_id=entorno.vendedor.id,
            timestamp=creado,
        )
    )
    db.add(
        HistorialEstado(
            solicitud_id=solicitud.id,
            de=Estado.EN_PROCESO,
            a=Estado.COTIZADA,
            usuario_id=entorno.vendedor.id,
            timestamp=creado + timedelta(hours=2),
        )
    )
    db.commit()

    ws = _hoja(client.get(EXPORT, headers=auth_headers(entorno.admin)))
    encabezados = [c.value for c in ws[1]]
    assert encabezados[:7] == [
        "Folio",
        "Cliente",
        "Sucursal",
        "Vendedor",
        "Comprador",
        "Estado",
        "Prioridad",
    ]
    fila = next(f for f in ws.iter_rows(min_row=2, values_only=True) if f[0] == "EXP-1")
    assert fila[1] == "DINCO"
    assert fila[2] == "Export Suc"
    assert fila[3] == entorno.vendedor.nombre
    assert fila[5] == "CONFIRMADA"
    # Fecha en la zona horaria de la sucursal (UTC-6): 21:00Z → 15:00 local.
    esperado = creado.astimezone(ZoneInfo(entorno.sucursal.timezone)).replace(tzinfo=None)
    assert fila[7] == esperado
    # Último ciclo cerrado: 2 horas hábiles (15:00→17:00 local de un jueves).
    assert fila[11] == "ESPERADA" and fila[12] == 2.0
    assert fila[13] == 1234.5 and fila[14] == "MXN"


def test_export_respeta_filtros_y_scoping(client, db, entorno, auth_headers, make_user):
    _enviada(client, entorno, auth_headers)  # del vendedor
    otro = make_user(Rol.VENDEDOR, sucursal_id=entorno.sucursal.id)
    headers_otro = auth_headers(otro)
    r = client.post(BASE, headers=headers_otro, json={"cliente": "DINCO", "partidas": [PARTIDA]})
    assert client.post(f"{BASE}/{r.json()['id']}/enviar", headers=headers_otro).status_code == 200

    # El vendedor exporta SOLO lo suyo (mismo scoping del listado).
    ws = _hoja(client.get(EXPORT, headers=auth_headers(entorno.vendedor)))
    assert ws.max_row == 2  # encabezado + 1

    # Filtro aplicado: estado sin filas → solo encabezado.
    ws = _hoja(
        client.get(EXPORT, params={"estado": "CANCELADA"}, headers=auth_headers(entorno.admin))
    )
    assert ws.max_row == 1


def test_export_limite_422(client, entorno, auth_headers, monkeypatch):
    from app.modules.metricas import export as export_mod

    monkeypatch.setattr(export_mod, "EXPORT_MAX_FILAS", 1)
    _enviada(client, entorno, auth_headers)
    _enviada(client, entorno, auth_headers)
    r = client.get(EXPORT, headers=auth_headers(entorno.admin))
    assert r.status_code == 422
    assert r.json()["code"] == "export_demasiado_grande"
