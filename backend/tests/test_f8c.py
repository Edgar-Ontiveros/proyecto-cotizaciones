"""F8c: moneda por renglón + tipo de cambio al confirmar, permisos v2 por
área (matriz de gestión exhaustiva + scoping), % no encontrados y la vista
del comprador sobre CONFIRMADA."""

from decimal import Decimal
from itertools import count
from types import SimpleNamespace

import pytest
from alembic.config import Config
from sqlalchemy import create_engine, text

from alembic import command
from app.models.sucursal import CompradorSucursal
from app.models.usuario import Rol
from app.modules.usuarios.service import MATRIZ_GESTION
from tests.conftest import BACKEND_DIR, _url, drop_database, recreate_database

BASE = "/api/v1/solicitudes"
USUARIOS = "/api/v1/usuarios"

PARTIDA_PZ = {"cantidad": "20", "unidad": "PZ", "descripcion": 'ANGULO 2" X 1/4"'}
PARTIDA_KG = {"cantidad": "100", "unidad": "KG", "descripcion": "SOLERA INOX 1/4 X 2"}

_emails = count(1)


def _email() -> str:
    return f"f8c{next(_emails)}@test.demo"


@pytest.fixture
def entorno(db, make_user, make_sucursal):
    suc_a = make_sucursal("F8c A")
    suc_b = make_sucursal("F8c B")
    comprador = make_user(Rol.COMPRADOR)
    db.add(CompradorSucursal(comprador_id=comprador.id, sucursal_id=suc_a.id, titular=True))
    db.commit()
    return SimpleNamespace(
        suc_a=suc_a,
        suc_b=suc_b,
        comprador=comprador,
        otro_comprador=make_user(Rol.COMPRADOR),
        vendedor=make_user(Rol.VENDEDOR, sucursal_id=suc_a.id),
        vendedor_b=make_user(Rol.VENDEDOR, sucursal_id=suc_b.id),
        gsuc=make_user(Rol.GERENTE_SUCURSAL, sucursal_id=suc_a.id),
        gcompras=make_user(Rol.GERENTE_COMPRAS),
        dventas=make_user(Rol.DIRECTOR_VENTAS),
        admin=make_user(Rol.ADMIN),
    )


@pytest.fixture
def cotizada_mixta(client, entorno, auth_headers):
    """2 partidas → opción A MIXTA: 20 PZ × 600 MXN = 12,000.00 MXN y
    100 KG × 5.00 USD = 500.00 USD (el ejemplo exacto del modelo F8c)."""
    headers_v = auth_headers(entorno.vendedor)
    r = client.post(
        BASE, headers=headers_v, json={"cliente": "DINCO", "partidas": [PARTIDA_PZ, PARTIDA_KG]}
    )
    sid = r.json()["id"]
    assert client.post(f"{BASE}/{sid}/enviar", headers=headers_v).status_code == 200
    detalle = client.get(f"{BASE}/{sid}", headers=headers_v).json()
    pid_pz, pid_kg = [p["id"] for p in detalle["partidas"]]
    r = client.put(
        f"{BASE}/{sid}/opciones/A",
        headers=auth_headers(entorno.comprador),
        json={
            "vigencia": "2026-08-31",
            "renglones": [
                {
                    "partida_id": pid_pz,
                    "moneda": "MXN",
                    "precio_unitario": "600.00",
                    "tiempo_entrega": "1 semana",
                    "proveedor": "Aceros del Norte",
                },
                {
                    "partida_id": pid_kg,
                    "moneda": "USD",
                    "precio_unitario": "5.00",
                    "tiempo_entrega": "3 semanas",
                    "proveedor": "Rolled Alloys",
                },
            ],
        },
    )
    assert r.status_code == 200, r.text
    assert (
        client.post(f"{BASE}/{sid}/cotizar", headers=auth_headers(entorno.comprador)).status_code
        == 200
    )
    return sid


# ------------------------------------------------- moneda por renglón + TC


def test_subtotales_mixtos_y_referencia_dual(client, entorno, cotizada_mixta, auth_headers):
    headers_v = auth_headers(entorno.vendedor)
    detalle = client.get(f"{BASE}/{cotizada_mixta}", headers=headers_v).json()
    opcion = detalle["opciones"][0]
    # Subtotales por moneda, JAMÁS sumados entre sí.
    assert opcion["total_mxn"] == "12000.00" and opcion["total_usd"] == "500.00"
    fila = next(
        i for i in client.get(BASE, headers=headers_v).json()["items"] if i["id"] == cotizada_mixta
    )
    assert fila["referencia_mxn"] == "12000.00" and fila["referencia_usd"] == "500.00"


def test_confirmar_mixta_exige_tc_y_consolida(client, db, entorno, cotizada_mixta, auth_headers):
    headers_v = auth_headers(entorno.vendedor)
    # Sin TC → 422 exacto.
    r = client.post(f"{BASE}/{cotizada_mixta}/seleccionar", headers=headers_v, json={"letra": "A"})
    assert r.status_code == 422 and r.json()["code"] == "tipo_cambio_requerido"
    # Con TC → consolidado EXACTO: 12,000 + 500 × 18.5000 = 21,250.00 MXN.
    r = client.post(
        f"{BASE}/{cotizada_mixta}/seleccionar",
        headers=headers_v,
        json={"letra": "A", "tipo_cambio": "18.5"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["monto_confirmado"] == "21250.00"
    assert body["moneda_confirmada"] == "MXN"
    assert body["tipo_cambio"] == "18.5000"


def test_tc_en_opcion_cien_por_ciento_mxn_422(client, entorno, auth_headers):
    headers_v = auth_headers(entorno.vendedor)
    r = client.post(BASE, headers=headers_v, json={"cliente": "DINCO", "partidas": [PARTIDA_PZ]})
    sid = r.json()["id"]
    assert client.post(f"{BASE}/{sid}/enviar", headers=headers_v).status_code == 200
    detalle = client.get(f"{BASE}/{sid}", headers=headers_v).json()
    r = client.put(
        f"{BASE}/{sid}/opciones/A",
        headers=auth_headers(entorno.comprador),
        json={
            "vigencia": "2026-08-31",
            "renglones": [
                {
                    "partida_id": detalle["partidas"][0]["id"],
                    "moneda": "MXN",
                    "precio_unitario": "10.00",
                    "tiempo_entrega": "1 semana",
                }
            ],
        },
    )
    assert r.status_code == 200
    assert (
        client.post(f"{BASE}/{sid}/cotizar", headers=auth_headers(entorno.comprador)).status_code
        == 200
    )
    r = client.post(
        f"{BASE}/{sid}/seleccionar", headers=headers_v, json={"letra": "A", "tipo_cambio": "18.5"}
    )
    assert r.status_code == 422 and r.json()["code"] == "tipo_cambio_invalido"
    # Sin TC → confirma sin datos basura.
    r = client.post(f"{BASE}/{sid}/seleccionar", headers=headers_v, json={"letra": "A"})
    assert r.status_code == 200 and r.json()["tipo_cambio"] is None


def test_kpi_confirmado_consolidado_y_export_con_tc(client, entorno, cotizada_mixta, auth_headers):
    headers_v = auth_headers(entorno.vendedor)
    r = client.post(
        f"{BASE}/{cotizada_mixta}/seleccionar",
        headers=headers_v,
        json={"letra": "A", "tipo_cambio": "18.5"},
    )
    assert r.status_code == 200
    # KPI: confirmado = UNA serie MXN consolidada; desglose original aparte.
    body = client.get("/api/v1/metricas/resumen", headers=auth_headers(entorno.admin)).json()
    assert body["dinero_confirmado"] == {"MXN": "21250.00"}
    assert body["dinero_confirmado_desglose"] == {"MXN": "12000.00", "USD": "500.00"}

    # Export: desglose + TC + consolidado.
    from io import BytesIO

    from openpyxl import load_workbook

    r = client.get("/api/v1/solicitudes/export", headers=auth_headers(entorno.admin))
    ws = load_workbook(BytesIO(r.content)).active
    encabezados = [c.value for c in ws[1]]
    assert encabezados[13:17] == ["Monto MXN", "Monto USD", "Tipo de cambio", "Confirmado MXN"]
    fila = next(f for f in ws.iter_rows(min_row=2, values_only=True) if f[5] == "CONFIRMADA")
    assert (fila[13], fila[14]) == (12000, 500)
    assert float(fila[15]) == 18.5 and fila[16] == 21250


# --------------------------------------------- fix 2a: comprador en CONFIRMADA


def test_comprador_ve_proveedor_en_confirmada(client, entorno, cotizada_mixta, auth_headers):
    r = client.post(
        f"{BASE}/{cotizada_mixta}/seleccionar",
        headers=auth_headers(entorno.vendedor),
        json={"letra": "A", "tipo_cambio": "18.5"},
    )
    assert r.status_code == 200
    # El comprador CONSERVA el proveedor tras la confirmación (lo necesita
    # para la orden de compra); la ganadora es identificable.
    detalle = client.get(f"{BASE}/{cotizada_mixta}", headers=auth_headers(entorno.comprador)).json()
    assert detalle["estado"] == "CONFIRMADA"
    assert detalle["opcion_seleccionada_id"] == detalle["opciones"][0]["id"]
    proveedores = {r_["proveedor"] for r_ in detalle["opciones"][0]["renglones"]}
    assert proveedores == {"Aceros del Norte", "Rolled Alloys"}
    # Todo el lado ventas sigue SIN la clave.
    for usuario in (entorno.vendedor, entorno.gsuc, entorno.dventas):
        detalle = client.get(f"{BASE}/{cotizada_mixta}", headers=auth_headers(usuario)).json()
        for renglon in detalle["opciones"][0]["renglones"]:
            assert "proveedor" not in renglon, usuario.rol


# ------------------------------------------------- permisos v2: matriz


def _payload(rol: Rol, sucursal_id: int | None) -> dict:
    datos: dict = {"nombre": f"Prueba {rol.value}", "email": _email(), "rol": rol.value}
    if sucursal_id is not None:
        datos["sucursal_id"] = sucursal_id
    return datos


def test_matriz_de_gestion_exhaustiva(client, db, entorno, make_user, auth_headers):
    """CADA gestor intenta crear y editar CADA rol: permitido según la
    MATRIZ_GESTION (el dato es la fuente de verdad) o 403 exacto."""
    gestores = {
        Rol.ADMIN: entorno.admin,
        Rol.DIRECTOR_VENTAS: entorno.dventas,
        Rol.GERENTE_SUCURSAL: entorno.gsuc,
        Rol.GERENTE_COMPRAS: entorno.gcompras,
    }
    sucursal_de = {
        Rol.VENDEDOR: entorno.suc_a.id,
        Rol.GERENTE_SUCURSAL: entorno.suc_a.id,
    }
    objetivos = {rol: make_user(rol, sucursal_id=sucursal_de.get(rol)) for rol in Rol}
    for rol_gestor, gestor in gestores.items():
        headers = auth_headers(gestor)
        permitidos = MATRIZ_GESTION[rol_gestor]
        for rol_objetivo in Rol:
            esperado_ok = rol_objetivo in permitidos
            # CREAR
            r = client.post(
                USUARIOS,
                headers=headers,
                json=_payload(rol_objetivo, sucursal_de.get(rol_objetivo)),
            )
            assert (r.status_code == 201) == esperado_ok, (
                f"crear {rol_gestor.value}→{rol_objetivo.value}: {r.status_code}"
            )
            if not esperado_ok:
                assert r.json()["code"] == "gestion_no_permitida"
            # EDITAR (nombre) sobre un usuario existente de ese rol
            r = client.patch(
                f"{USUARIOS}/{objetivos[rol_objetivo].id}",
                headers=headers,
                json={"nombre": f"Editado por {rol_gestor.value}"},
            )
            assert (r.status_code == 200) == esperado_ok, (
                f"editar {rol_gestor.value}→{rol_objetivo.value}: {r.status_code}"
            )

    # Alcance de sucursal: el gerente NO toca vendedores de otra sucursal.
    r = client.post(
        USUARIOS, headers=auth_headers(entorno.gsuc), json=_payload(Rol.VENDEDOR, entorno.suc_b.id)
    )
    assert r.status_code == 403 and r.json()["code"] == "gestion_no_permitida"
    r = client.patch(
        f"{USUARIOS}/{entorno.vendedor_b.id}",
        headers=auth_headers(entorno.gsuc),
        json={"nombre": "X"},
    )
    assert r.status_code == 403
    # El director de ventas SÍ mueve vendedores entre sucursales.
    r = client.patch(
        f"{USUARIOS}/{entorno.vendedor_b.id}",
        headers=auth_headers(entorno.dventas),
        json={"sucursal_id": entorno.suc_a.id},
    )
    assert r.status_code == 200 and r.json()["sucursal_id"] == entorno.suc_a.id


def test_nadie_se_cambia_rol_ni_activo_a_si_mismo(client, entorno, auth_headers):
    # Generalizada (F8c) para TODOS los gestores, no solo admin.
    for gestor in (entorno.admin, entorno.dventas, entorno.gcompras, entorno.gsuc):
        headers = auth_headers(gestor)
        r = client.patch(
            f"{USUARIOS}/{gestor.id}",
            headers=headers,
            json={"rol": "vendedor", "sucursal_id": entorno.suc_a.id},
        )
        assert r.status_code == 422 and r.json()["code"] == "no_auto_degradacion", gestor.rol
        r = client.post(f"{USUARIOS}/{gestor.id}/desactivar", headers=headers)
        assert r.status_code in (400, 403), gestor.rol
        if r.status_code == 400:
            assert r.json()["code"] == "no_auto_desactivacion"


def test_gestion_de_listado_scoped(client, entorno, make_user, auth_headers):
    make_user(Rol.VENDEDOR, sucursal_id=entorno.suc_b.id)
    # gerente_sucursal: SOLO vendedores de su sucursal.
    items = client.get(USUARIOS, headers=auth_headers(entorno.gsuc)).json()["items"]
    assert items and all(
        u["rol"] == "vendedor" and u["sucursal_id"] == entorno.suc_a.id for u in items
    )
    # gerente_compras: SOLO compradores.
    items = client.get(USUARIOS, headers=auth_headers(entorno.gcompras)).json()["items"]
    assert items and all(u["rol"] == "comprador" for u in items)
    # director_ventas: vendedores (todas las sucursales) y gerentes_sucursal.
    items = client.get(USUARIOS, headers=auth_headers(entorno.dventas)).json()["items"]
    roles = {u["rol"] for u in items}
    assert roles == {"vendedor", "gerente_sucursal"}


# ------------------------------------------------- permisos v2: scoping/área


def test_director_ventas_global_sin_proveedor_y_confirma(
    client, entorno, cotizada_mixta, auth_headers
):
    headers = auth_headers(entorno.dventas)
    # Ve la solicitud (global) SIN clave proveedor.
    detalle = client.get(f"{BASE}/{cotizada_mixta}", headers=headers).json()
    for renglon in detalle["opciones"][0]["renglones"]:
        assert "proveedor" not in renglon
    # Confirma CUALQUIER solicitud y el historial LO registra a él.
    r = client.post(
        f"{BASE}/{cotizada_mixta}/seleccionar",
        headers=headers,
        json={"letra": "A", "tipo_cambio": "18.5"},
    )
    assert r.status_code == 200, r.text
    detalle = client.get(f"{BASE}/{cotizada_mixta}", headers=headers).json()
    confirmacion = next(h for h in detalle["historial"] if h["a"] == "CONFIRMADA")
    assert confirmacion["usuario_id"] == entorno.dventas.id


def test_gerente_compras_ve_con_proveedor_pero_no_cotiza(
    client, entorno, cotizada_mixta, auth_headers
):
    headers = auth_headers(entorno.gcompras)
    detalle = client.get(f"{BASE}/{cotizada_mixta}", headers=headers).json()
    proveedores = {r_["proveedor"] for r_ in detalle["opciones"][0]["renglones"]}
    assert "Aceros del Norte" in proveedores  # CON proveedor (área compras)
    # Pero NO captura/cotiza/rechaza ni edita (reasigna, no cotiza).
    r = client.put(f"{BASE}/{cotizada_mixta}/opciones/B", headers=headers, json={"renglones": []})
    assert r.status_code == 403
    assert client.post(f"{BASE}/{cotizada_mixta}/cotizar", headers=headers).status_code == 403
    # Métricas por vendedor: 403 explícito.
    assert client.get("/api/v1/metricas/por-vendedor", headers=headers).status_code == 403
    # Métricas por comprador: SÍ.
    assert client.get("/api/v1/metricas/por-comprador", headers=headers).status_code == 200
    # El lado ventas gerencial al revés:
    assert (
        client.get(
            "/api/v1/metricas/por-comprador", headers=auth_headers(entorno.dventas)
        ).status_code
        == 403
    )
    assert (
        client.get("/api/v1/metricas/por-comprador", headers=auth_headers(entorno.gsuc)).status_code
        == 403
    )


def test_gerente_sucursal_fuera_de_su_sucursal_404(client, entorno, auth_headers, make_sucursal):
    headers_v = auth_headers(entorno.vendedor_b)
    r = client.post(BASE, headers=headers_v, json={"cliente": "AJENA", "partidas": [PARTIDA_PZ]})
    sid = r.json()["id"]
    # BORRADOR de otra sucursal: invisible para el gerente de la sucursal A.
    assert client.get(f"{BASE}/{sid}", headers=auth_headers(entorno.gsuc)).status_code == 404


def test_reasignaciones_por_area(client, db, entorno, auth_headers):
    headers_v = auth_headers(entorno.vendedor)
    r = client.post(BASE, headers=headers_v, json={"cliente": "DINCO", "partidas": [PARTIDA_PZ]})
    sid = r.json()["id"]
    assert client.post(f"{BASE}/{sid}/enviar", headers=headers_v).status_code == 200
    # gerente_compras reasigna compradores…
    r = client.post(
        f"{BASE}/{sid}/reasignar-comprador",
        headers=auth_headers(entorno.gcompras),
        json={"comprador_id": entorno.otro_comprador.id},
    )
    assert r.status_code == 200
    # …pero NO vendedores; y el lado ventas al revés.
    r = client.post(
        f"{BASE}/{sid}/reasignar-vendedor",
        headers=auth_headers(entorno.gcompras),
        json={"vendedor_id": entorno.vendedor.id},
    )
    assert r.status_code == 403
    r = client.post(
        f"{BASE}/{sid}/reasignar-comprador",
        headers=auth_headers(entorno.dventas),
        json={"comprador_id": entorno.comprador.id},
    )
    assert r.status_code == 403


# ------------------------------------------------- % de renglones no encontrados


def test_pct_no_encontrados(client, entorno, auth_headers, make_user, db):
    headers_v = auth_headers(entorno.vendedor)
    r = client.post(
        BASE, headers=headers_v, json={"cliente": "DINCO", "partidas": [PARTIDA_PZ, PARTIDA_KG]}
    )
    sid = r.json()["id"]
    assert client.post(f"{BASE}/{sid}/enviar", headers=headers_v).status_code == 200
    detalle = client.get(f"{BASE}/{sid}", headers=headers_v).json()
    pid_1, pid_2 = [p["id"] for p in detalle["partidas"]]
    r = client.put(
        f"{BASE}/{sid}/opciones/A",
        headers=auth_headers(entorno.comprador),
        json={
            "vigencia": "2026-08-31",
            "renglones": [
                {
                    "partida_id": pid_1,
                    "moneda": "MXN",
                    "precio_unitario": "10.00",
                    "tiempo_entrega": "1 semana",
                },
                {"partida_id": pid_2, "no_encontrada": True},
            ],
        },
    )
    assert r.status_code == 200, r.text

    body = client.get(
        "/api/v1/metricas/no-encontrados", headers=auth_headers(entorno.gcompras)
    ).json()
    assert body["total_renglones"] == 2 and body["no_encontrados"] == 1
    assert body["pct"] == 0.5
    fila = next(g for g in body["por_comprador"] if g["id"] == entorno.comprador.id)
    assert fila["no_encontrados"] == 1 and fila["pct"] == 0.5
    assert body["top_materiales"][0]["valor"] == "SOLERA INOX 1/4 X 2"

    # Visible SOLO para compras global y admin.
    assert (
        client.get(
            "/api/v1/metricas/no-encontrados", headers=auth_headers(entorno.admin)
        ).status_code
        == 200
    )
    for usuario in (entorno.dventas, entorno.gsuc, entorno.vendedor, entorno.comprador):
        r = client.get("/api/v1/metricas/no-encontrados", headers=auth_headers(usuario))
        assert r.status_code == 403, usuario.rol


# ------------------------------------------------- migración de datos


def test_migracion_f8c_roles_y_consolidado_usd():
    """Upgrade: 'gerente' → 'gerente_sucursal'; la moneda baja al renglón; la
    confirmada USD se consolida con TC 18.50. Downgrade: regresa roles y
    revierte el TC demo."""
    db_name = f"{_url.database}_migraf8c"
    recreate_database(db_name)
    try:
        cfg = Config(str(BACKEND_DIR / "alembic.ini"))
        cfg.set_main_option("script_location", str(BACKEND_DIR / "alembic"))
        cfg.set_main_option(
            "sqlalchemy.url", _url.set(database=db_name).render_as_string(hide_password=False)
        )
        command.upgrade(cfg, "923c7cfecbc4")  # hasta ANTES de F8c
        engine = create_engine(_url.set(database=db_name))
        try:
            with engine.begin() as conn:
                conn.execute(
                    text(
                        "INSERT INTO usuarios (nombre, email, password_hash, rol, activo,"
                        " must_change_password) VALUES ('G', 'g@x.demo', 'h', 'gerente', true,"
                        " false), ('V', 'v@x.demo', 'h', 'vendedor', true, false)"
                    )
                )
                conn.execute(
                    text(
                        "INSERT INTO sucursales (nombre, prefijo_folio, timezone, activa)"
                        " VALUES ('S', 'SSS', 'America/Chihuahua', true)"
                    )
                )
                conn.execute(
                    text(
                        "INSERT INTO solicitudes (vendedor_id, sucursal_id, estado, prioridad,"
                        " monto_confirmado, moneda_confirmada) VALUES"
                        " (2, 1, 'CONFIRMADA', 'NORMAL', 351.00, 'USD')"
                    )
                )
                conn.execute(
                    text(
                        "INSERT INTO solicitudes (vendedor_id, sucursal_id, estado, prioridad)"
                        " VALUES (2, 1, 'EN_PROCESO', 'NORMAL')"
                    )
                )
                conn.execute(
                    text(
                        "INSERT INTO solicitud_partidas (solicitud_id, num_partida, cantidad,"
                        " unidad, descripcion) VALUES (2, 1, 20, 'PZ', 'ANGULO')"
                    )
                )
                conn.execute(
                    text(
                        "INSERT INTO cotizacion_opciones (solicitud_id, letra, moneda, total,"
                        " completa) VALUES (2, 'A', 'USD', 100.00, false)"
                    )
                )
                conn.execute(
                    text(
                        "INSERT INTO opcion_partidas (opcion_id, partida_id, cantidad, unidad,"
                        " precio_unitario, importe) VALUES (1, 1, 20, 'PZ', 5.00, 100.00)"
                    )
                )
            command.upgrade(cfg, "head")
            with engine.connect() as conn:
                assert (
                    conn.execute(
                        text("SELECT rol FROM usuarios WHERE email = 'g@x.demo'")
                    ).scalar_one()
                    == "gerente_sucursal"
                )
                fila = conn.execute(
                    text(
                        "SELECT monto_confirmado, moneda_confirmada, tipo_cambio FROM"
                        " solicitudes WHERE estado = 'CONFIRMADA'"
                    )
                ).one()
                # 351.00 USD × 18.50 = 6,493.50 consolidado MXN.
                assert fila == (Decimal("6493.50"), "MXN", Decimal("18.5000"))
                moneda, mxn, usd = conn.execute(
                    text(
                        "SELECT op.moneda, co.total_mxn, co.total_usd FROM opcion_partidas op"
                        " JOIN cotizacion_opciones co ON co.id = op.opcion_id"
                    )
                ).one()
                # La moneda bajó de la opción al renglón; el total al subtotal.
                assert (moneda, mxn, usd) == ("USD", Decimal("0.00"), Decimal("100.00"))

            command.downgrade(cfg, "923c7cfecbc4")
            with engine.connect() as conn:
                assert (
                    conn.execute(
                        text("SELECT rol FROM usuarios WHERE email = 'g@x.demo'")
                    ).scalar_one()
                    == "gerente"
                )
                fila = conn.execute(
                    text(
                        "SELECT monto_confirmado, moneda_confirmada FROM solicitudes"
                        " WHERE estado = 'CONFIRMADA'"
                    )
                ).one()
                assert fila == (Decimal("351.00"), "USD")  # TC demo revertido
        finally:
            engine.dispose()
    finally:
        drop_database(db_name)
