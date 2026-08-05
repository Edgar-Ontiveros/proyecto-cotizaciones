"""F10: ajustes de producción v1.

Punto 1 — regresión del flujo completo reportado en producción:
cotizada con 2 opciones → cambio solicitado → aprobado con ajustes →
el vendedor DEBE poder seleccionar cualquier opción → CONFIRMADA.

Escenario (aritmética heredada de test_f8h):
- P1: 20 PZ → cambio a 500 KG (la unidad cambia: precios de P1 se reponen
  con ajustes 94.80 MXN/KG en A y 3.10 USD/KG en B).
- Tras aprobar: A total_mxn 48,400.00; B 500.00 MXN + 1,550.00 USD, TC 18.5
  → consolidado B = 500 + 1,550×18.5 = 29,175.00.
"""

from tests.test_f8h import (  # noqa: F401  (fixture entorno re-exportada)
    BASE,
    CAMBIOS,
    PARTIDA1,
    PARTIDA2,
    _cambio_pendiente,
    entorno,
)


def _aprobar_con_ajustes(client, auth_headers, entorno, cambio_id, p1):
    return client.post(
        f"{CAMBIOS}/{cambio_id}/aprobar",
        headers=auth_headers(entorno.comprador),
        json={
            "ajustes": [
                {"opcion_letra": "A", "partida_id": p1, "precio_unitario": "94.80"},
                {"opcion_letra": "B", "partida_id": p1, "precio_unitario": "3.10"},
            ]
        },
    )


def test_flujo_cambiar_aprobar_seleccionar_confirmada(
    client, entorno, auth_headers, con_comprobante
):
    """El flujo del reporte de producción, de punta a punta."""
    sid, p1, _p2, cambio_id = _cambio_pendiente(client, entorno, auth_headers)
    r = _aprobar_con_ajustes(client, auth_headers, entorno, cambio_id, p1)
    assert r.status_code == 200, r.text

    # El detalle del vendedor ya no reporta cambio pendiente.
    detalle = client.get(f"{BASE}/{sid}", headers=auth_headers(entorno.vendedor)).json()
    assert detalle["cambio_pendiente"] is False

    # Comprobante y selección de la opción B (la mixta, la más frágil).
    con_comprobante(sid, entorno.vendedor)
    r = client.post(
        f"{BASE}/{sid}/seleccionar",
        headers=auth_headers(entorno.vendedor),
        json={"letra": "B"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["estado"] == "CONFIRMADA"


def test_flujo_cambio_aprobado_seleccionar_opcion_a(client, entorno, auth_headers, con_comprobante):
    """Misma base pero seleccionando la opción A (100% MXN)."""
    sid, p1, _p2, cambio_id = _cambio_pendiente(client, entorno, auth_headers)
    assert _aprobar_con_ajustes(client, auth_headers, entorno, cambio_id, p1).status_code == 200
    con_comprobante(sid, entorno.vendedor)
    r = client.post(
        f"{BASE}/{sid}/seleccionar",
        headers=auth_headers(entorno.vendedor),
        json={"letra": "A"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["estado"] == "CONFIRMADA"
    # El monto confirmado NO existe en el JSON del vendedor (F8e): se
    # verifica con el detalle del admin. A = 500 × 94.80 + 1,000 = 48,400.
    detalle = client.get(f"{BASE}/{sid}", headers=auth_headers(entorno.admin)).json()
    assert detalle["monto_confirmado"] == "48400.00"


# ---------------------------------------------- P3: paridad gcompras=comprador


def test_paridad_detalle_gerente_compras_igual_comprador(
    client, entorno, auth_headers, con_comprobante
):
    """F10 p.3: gerente_compras recibe EXACTAMENTE el mismo JSON de detalle
    que el comprador asignado en ENVIADA, EN_PROCESO, COTIZADA y CONFIRMADA."""
    headers_v = auth_headers(entorno.vendedor)
    headers_c = auth_headers(entorno.comprador)
    headers_g = auth_headers(entorno.gerente_compras)

    def paridad(sid: int, estado: str) -> None:
        d_comprador = client.get(f"{BASE}/{sid}", headers=headers_c).json()
        d_gcompras = client.get(f"{BASE}/{sid}", headers=headers_g).json()
        assert d_comprador["estado"] == estado
        assert d_comprador == d_gcompras, f"difieren en {estado}"

    r = client.post(
        BASE, headers=headers_v, json={"cliente": "DINCO", "partidas": [PARTIDA1, PARTIDA2]}
    )
    sid = r.json()["id"]
    assert client.post(f"{BASE}/{sid}/enviar", headers=headers_v).status_code == 200
    paridad(sid, "ENVIADA")

    assert client.post(f"{BASE}/{sid}/tomar", headers=headers_c).status_code == 200
    paridad(sid, "EN_PROCESO")

    detalle = client.get(f"{BASE}/{sid}", headers=headers_c).json()
    p1, p2 = [p["id"] for p in detalle["partidas"]]
    r = client.put(
        f"{BASE}/{sid}/opciones/A",
        headers=headers_c,
        json={
            "vigencia": "2026-09-30",
            "renglones": [
                {
                    "partida_id": p1,
                    "moneda": "MXN",
                    "precio_unitario": "250.00",
                    "tiempo_entrega": "1 semana",
                    "proveedor": "Aceros del Norte",
                },
                {
                    "partida_id": p2,
                    "moneda": "MXN",
                    "precio_unitario": "100.00",
                    "tiempo_entrega": "2 semanas",
                },
            ],
        },
    )
    assert r.status_code == 200, r.text
    assert client.post(f"{BASE}/{sid}/cotizar", headers=headers_c, json={}).status_code == 200
    paridad(sid, "COTIZADA")

    con_comprobante(sid, entorno.vendedor)
    r = client.post(f"{BASE}/{sid}/seleccionar", headers=headers_v, json={"letra": "A"})
    assert r.status_code == 200, r.text
    paridad(sid, "CONFIRMADA")


# ------------------------------------------------- P6: comprobantes múltiples


def _pdf():
    from io import BytesIO

    from app.modules.archivos.service import pdf_minimo

    return BytesIO(pdf_minimo())


def _subir_comprobante(client, headers, sid, nombre="c.pdf"):
    return client.post(
        f"{BASE}/{sid}/comprobante",
        headers=headers,
        files={"archivo": (nombre, _pdf(), "application/pdf")},
    )


def _cotizada_simple(client, entorno, auth_headers):
    sid, _p1, _p2 = _cotizada_mixta_import(client, entorno, auth_headers)
    return sid


from tests.test_f8h import _cotizada_mixta as _cotizada_mixta_import  # noqa: E402


def test_confirmar_exige_al_menos_uno_y_con_dos_pasa(client, entorno, auth_headers):
    """F10 p.6: con 0 comprobantes → 422; con 2 → CONFIRMADA."""
    sid = _cotizada_simple(client, entorno, auth_headers)
    headers_v = auth_headers(entorno.vendedor)
    r = client.post(f"{BASE}/{sid}/seleccionar", headers=headers_v, json={"letra": "A"})
    assert r.status_code == 422 and r.json()["code"] == "comprobante_requerido"
    assert _subir_comprobante(client, headers_v, sid, "anticipo.pdf").status_code == 200
    assert _subir_comprobante(client, headers_v, sid, "orden.pdf").status_code == 200
    r = client.post(f"{BASE}/{sid}/seleccionar", headers=headers_v, json={"letra": "A"})
    assert r.status_code == 200 and r.json()["estado"] == "CONFIRMADA"
    detalle = client.get(f"{BASE}/{sid}", headers=headers_v).json()
    assert len(detalle["comprobantes"]) == 2


def test_eliminar_solo_subidor_o_admin_y_solo_antes_de_confirmar(client, entorno, auth_headers):
    """Eliminar individual: quien lo subió o admin; el gerente de la sucursal
    NO borra el del vendedor. Tras CONFIRMADA, 409 para todos."""
    sid = _cotizada_simple(client, entorno, auth_headers)
    headers_v = auth_headers(entorno.vendedor)
    a1 = _subir_comprobante(client, headers_v, sid, "a1.pdf").json()
    a2 = _subir_comprobante(client, headers_v, sid, "a2.pdf").json()
    a3 = _subir_comprobante(client, auth_headers(entorno.gerente), sid, "a3.pdf").json()

    # El gerente subió a3 pero NO puede borrar a1 (del vendedor) → 403.
    r = client.delete(
        f"{BASE}/{sid}/comprobantes/{a1['id']}", headers=auth_headers(entorno.gerente)
    )
    assert r.status_code == 403
    # El vendedor borra el suyo; el admin borra el del gerente.
    from app.modules.archivos.service import ruta_de

    assert (
        client.delete(f"{BASE}/{sid}/comprobantes/{a1['id']}", headers=headers_v).status_code == 204
    )
    assert not ruta_de(a1["id"]).exists()  # borrado seguro del disco
    assert (
        client.delete(
            f"{BASE}/{sid}/comprobantes/{a3['id']}", headers=auth_headers(entorno.admin)
        ).status_code
        == 204
    )
    detalle = client.get(f"{BASE}/{sid}", headers=headers_v).json()
    assert [c["nombre_original"] for c in detalle["comprobantes"]] == ["a2.pdf"]
    eventos = [h["comentario"] for h in detalle["historial"] if h["de"] == h["a"]]
    assert "Comprobante eliminado (a1.pdf)" in eventos

    # Confirmada → subir y borrar dan 409 comprobante_inmutable.
    r = client.post(f"{BASE}/{sid}/seleccionar", headers=headers_v, json={"letra": "A"})
    assert r.status_code == 200
    r = client.delete(f"{BASE}/{sid}/comprobantes/{a2['id']}", headers=headers_v)
    assert r.status_code == 409 and r.json()["code"] == "comprobante_inmutable"
    r = _subir_comprobante(client, headers_v, sid, "tarde.pdf")
    assert r.status_code == 409 and r.json()["code"] == "comprobante_inmutable"


def test_esquema_archivos_sin_unique(db):
    """La migración 2db348b42524 quitó el UNIQUE(solicitud, tipo) y dejó un
    índice normal — el esquema migrado debe reflejarlo (up/down probado
    también contra la BD de desarrollo)."""
    from sqlalchemy import inspect

    inspector = inspect(db.get_bind())
    unicos = {u["name"] for u in inspector.get_unique_constraints("archivos")}
    assert "uq_archivos_solicitud_id" not in unicos
    indices = {i["name"]: i for i in inspector.get_indexes("archivos")}
    assert "ix_archivos_solicitud_tipo" in indices
    assert indices["ix_archivos_solicitud_tipo"]["unique"] is False


# ------------------------------------- P7: notificaciones, badge, filtro, export


def _campana(client, auth_headers, usuario):
    r = client.get("/api/v1/notificaciones", headers=auth_headers(usuario))
    return [(n["tipo"], n["mensaje"]) for n in r.json()["items"]]


def test_notifica_al_solicitar_a_comprador_y_gerentes_compras(client, entorno, auth_headers):
    """F10 p.7b: al SOLICITAR suenan la campana del comprador ASIGNADO y la de
    TODOS los gerentes de compras activos (en la práctica ellos operan el lado
    compras, F8c.1 — antes no recibían nada, causa raíz del reporte 7a)."""
    _sid, _p1, _p2, _cid = _cambio_pendiente(client, entorno, auth_headers)
    for usuario in (entorno.comprador, entorno.gerente_compras):
        tipos = [t for t, _ in _campana(client, auth_headers, usuario)]
        assert "cambio_solicitado" in tipos, usuario.rol
    # El otro comprador (no asignado) NO recibe nada del cambio.
    tipos_otro = [t for t, _ in _campana(client, auth_headers, entorno.otro_comprador)]
    assert "cambio_solicitado" not in tipos_otro


def test_notifica_al_resolver_distinguiendo_desenlaces(client, entorno, auth_headers):
    """Al RESOLVER, el solicitante distingue: aprobado-con-ajuste / aprobado /
    rechazado."""
    from tests.test_f8h import _cotizada_mixta

    # Caso 1: aprobado CON ajuste de precio.
    _sid, p1, _p2, cambio_id = _cambio_pendiente(client, entorno, auth_headers)
    assert _aprobar_con_ajustes(client, auth_headers, entorno, cambio_id, p1).status_code == 200
    mensajes = [
        m for t, m in _campana(client, auth_headers, entorno.vendedor) if t == "cambio_aprobado"
    ]
    assert any("ajustó el precio" in m for m in mensajes)

    # Caso 2: rechazado (con motivo).
    _sid2, _p1b, _p2b, cambio2 = _cambio_pendiente(client, entorno, auth_headers)
    r = client.post(
        f"{CAMBIOS}/{cambio2}/rechazar",
        headers=auth_headers(entorno.comprador),
        json={"comentario": "El proveedor no maneja KG"},
    )
    assert r.status_code == 200
    tipos = [t for t, _ in _campana(client, auth_headers, entorno.vendedor)]
    assert "cambio_rechazado" in tipos

    # Caso 3: aprobado SIN ajuste (solo cantidad, la unidad no cambia).
    sid3, p1c, _p2c = _cotizada_mixta(client, entorno, auth_headers)
    r = client.post(
        f"{BASE}/{sid3}/cambios",
        headers=auth_headers(entorno.vendedor),
        json={"comentario": None, "partidas": [{"partida_id": p1c, "cantidad_nueva": "25"}]},
    )
    assert r.status_code == 201
    r = client.post(
        f"{CAMBIOS}/{r.json()['id']}/aprobar",
        headers=auth_headers(entorno.comprador),
        json={"ajustes": []},
    )
    assert r.status_code == 200, r.text
    aprobados = [
        m for t, m in _campana(client, auth_headers, entorno.vendedor) if t == "cambio_aprobado"
    ]
    assert any("ajustó el precio" not in m for m in aprobados)


def test_filtro_cambio_pendiente_en_listado_y_export(client, entorno, auth_headers):
    """F10 p.7b: filtro cambio_pendiente=true en el listado y columna 'Cambio
    pendiente' en el export con el MISMO filtro."""
    from io import BytesIO

    from openpyxl import load_workbook

    from tests.test_f8h import _cotizada_mixta

    sid_con, _p1, _p2, _cid = _cambio_pendiente(client, entorno, auth_headers)
    sid_sin, _a, _b = _cotizada_mixta(client, entorno, auth_headers)

    headers = auth_headers(entorno.admin)
    todas = client.get(f"{BASE}", headers=headers).json()
    assert {s["id"] for s in todas["items"]} >= {sid_con, sid_sin}
    con = client.get(f"{BASE}?cambio_pendiente=true", headers=headers).json()
    assert {s["id"] for s in con["items"]} == {sid_con}
    assert con["items"][0]["cambio_pendiente"] is True

    r = client.get(f"{BASE}/export?cambio_pendiente=true", headers=headers)
    assert r.status_code == 200
    ws = load_workbook(BytesIO(r.content)).active
    encabezados = [c.value for c in ws[1]]
    assert "Cambio pendiente" in encabezados
    col = encabezados.index("Cambio pendiente") + 1
    assert ws.max_row == 2  # solo la que tiene cambio pendiente
    assert ws.cell(row=2, column=col).value == "Sí"


# --------------------------- F10.1 p.2b: badge verde CAMBIO APROBADO derivado


def test_cambio_aprobado_derivado_en_listado_y_detalle(
    client, entorno, auth_headers, con_comprobante
):
    """El campo cambio_aprobado es True SOLO mientras: último cambio APROBADO
    y estado COTIZADA. Al confirmar muere solo. El rechazado no lo enciende."""
    # Aprobado → True (listado de admin Y de vendedor; detalle también).
    sid, p1, _p2, cambio_id = _cambio_pendiente(client, entorno, auth_headers)
    assert _aprobar_con_ajustes(client, auth_headers, entorno, cambio_id, p1).status_code == 200
    for usuario in (entorno.admin, entorno.vendedor):
        items = client.get(f"{BASE}", headers=auth_headers(usuario)).json()["items"]
        assert next(s for s in items if s["id"] == sid)["cambio_aprobado"] is True, usuario.rol
    detalle = client.get(f"{BASE}/{sid}", headers=auth_headers(entorno.comprador)).json()
    assert detalle["cambio_aprobado"] is True

    # Confirmada → el badge muere solo (estado ya no es COTIZADA).
    con_comprobante(sid, entorno.vendedor)
    r = client.post(
        f"{BASE}/{sid}/seleccionar", headers=auth_headers(entorno.vendedor), json={"letra": "A"}
    )
    assert r.status_code == 200
    items = client.get(f"{BASE}", headers=auth_headers(entorno.admin)).json()["items"]
    assert next(s for s in items if s["id"] == sid)["cambio_aprobado"] is False

    # Rechazado → False (solo el aprobado lleva badge).
    sid2, _p1b, _p2b, cambio2 = _cambio_pendiente(client, entorno, auth_headers)
    r = client.post(
        f"{CAMBIOS}/{cambio2}/rechazar",
        headers=auth_headers(entorno.comprador),
        json={"comentario": "no procede"},
    )
    assert r.status_code == 200
    items = client.get(f"{BASE}", headers=auth_headers(entorno.admin)).json()["items"]
    assert next(s for s in items if s["id"] == sid2)["cambio_aprobado"] is False


def test_ultimo_cambio_manda_no_cualquier_aprobado(client, entorno, auth_headers):
    """Con historial aprobado→rechazado, manda el ÚLTIMO (rechazado): False."""
    sid, p1, _p2, cambio_id = _cambio_pendiente(client, entorno, auth_headers)
    assert _aprobar_con_ajustes(client, auth_headers, entorno, cambio_id, p1).status_code == 200
    # Segundo cambio sobre la misma solicitud, ahora rechazado.
    r = client.post(
        f"{BASE}/{sid}/cambios",
        headers=auth_headers(entorno.vendedor),
        json={"comentario": None, "partidas": [{"partida_id": p1, "cantidad_nueva": "600"}]},
    )
    assert r.status_code == 201
    r = client.post(
        f"{CAMBIOS}/{r.json()['id']}/rechazar",
        headers=auth_headers(entorno.comprador),
        json={"comentario": "ya no"},
    )
    assert r.status_code == 200
    items = client.get(f"{BASE}", headers=auth_headers(entorno.admin)).json()["items"]
    assert next(s for s in items if s["id"] == sid)["cambio_aprobado"] is False


# ----------------------- F10.2 p.4: orden por confirmado_en para Confirmadas


def test_listado_orden_confirmado_en(client, entorno, auth_headers, con_comprobante):
    """orden=confirmado_en regresa las CONFIRMADAS por fecha de confirmación
    DESC (la segunda confirmada aparece primero aunque se creó después)."""
    from tests.test_f8h import _cotizada_mixta

    sids = []
    for _ in range(2):
        sid, _p1, _p2 = _cotizada_mixta(client, entorno, auth_headers)
        con_comprobante(sid, entorno.vendedor)
        r = client.post(
            f"{BASE}/{sid}/seleccionar",
            headers=auth_headers(entorno.vendedor),
            json={"letra": "A"},
        )
        assert r.status_code == 200
        sids.append(sid)

    r = client.get(
        f"{BASE}?estado=CONFIRMADA&orden=confirmado_en",
        headers=auth_headers(entorno.comprador),
    )
    items = r.json()["items"]
    fechas = [s["confirmado_en"] for s in items]
    assert fechas == sorted(fechas, reverse=True)
    assert {s["id"] for s in items} >= set(sids)
    # Valor de orden inválido → 422 de validación, no un 500.
    r = client.get(f"{BASE}?orden=otra_cosa", headers=auth_headers(entorno.comprador))
    assert r.status_code == 422
