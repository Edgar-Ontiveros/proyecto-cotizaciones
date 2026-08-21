"""Export a Excel del listado de solicitudes (F6, §6).

Respeta EXACTAMENTE los mismos filtros y scoping del listado
(solicitudes/service.stmt_listado). Fechas en la zona horaria de la sucursal
de cada fila. Máximo 10,000 filas → 422 pidiendo filtrar más."""

from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.errors import AppError
from app.core.horario_habil import Banda
from app.core.permissions import get_current_user
from app.models.catalogos import MotivoRechazo
from app.models.cotizacion import CotizacionOpcion, Moneda
from app.models.historial import HistorialEstado
from app.models.solicitud import Estado, Prioridad, Solicitud
from app.models.sucursal import Sucursal
from app.models.usuario import Rol, Usuario
from app.modules.cotizaciones import service as cotizaciones_service
from app.modules.metricas import service as metricas_service
from app.modules.metricas.ciclos import cargar_ciclos
from app.modules.metricas.schemas import GrupoOut
from app.modules.metricas.service import Dimension, Filtros
from app.modules.metricas.tiempos import cargar_tiempos
from app.modules.solicitudes import service as solicitudes_service

router = APIRouter(tags=["export"])

EXPORT_MAX_FILAS = 10_000

# F14 §0a (regla de dinero, §4.9): estas columnas NO EXISTEN en el archivo
# del rol vendedor — ni encabezado ni celda, igual que en sus schemas JSON.
_COLUMNAS_SOLO_CONSOLIDADO = {"Tipo de cambio", "Confirmado MXN"}

_ENCABEZADOS = [
    ("Folio", 12),
    ("Cliente", 28),
    ("Sucursal", 16),
    ("Vendedor", 26),
    ("Comprador", 26),
    ("Estado", 14),
    ("Prioridad", 10),
    ("Proyecto", 10),
    ("Cambio pendiente", 14),
    ("Creado", 17),
    ("Enviado", 17),
    ("Cotizado", 17),
    ("Confirmado", 17),
    ("Banda último ciclo", 16),
    ("Horas hábiles último ciclo", 22),
    ("Total general (hrs naturales)", 24),
    ("Tiempo compras (hrs hábiles)", 24),
    ("Tiempo ventas (hrs hábiles)", 24),
    ("Monto MXN", 14),
    ("Monto USD", 14),
    ("Tipo de cambio", 12),
    ("Confirmado MXN", 16),
    ("Motivo", 34),
]
_FORMATO_FECHA = "yyyy-mm-dd hh:mm"


def _nombres_usuarios(db: Session, ids: set[int]) -> dict[int, str]:
    if not ids:
        return {}
    # .all() antes de dict(): dict() trata al Result como mapping (tiene
    # .keys()) e intenta indexarlo; .tuples() solo aporta el tipado.
    return dict(
        db.execute(select(Usuario.id, Usuario.nombre).where(Usuario.id.in_(ids))).tuples().all()
    )


def _sucursales(db: Session, ids: set[int]) -> dict[int, tuple[str, str]]:
    filas = db.execute(
        select(Sucursal.id, Sucursal.nombre, Sucursal.timezone).where(Sucursal.id.in_(ids))
    ).all()
    return {sid: (nombre, tz) for sid, nombre, tz in filas}


def _opciones_a(db: Session, ids: list[int]) -> dict[int, tuple[Decimal, Decimal]]:
    """(total_mxn, total_usd) de la opción A (referencia de una COTIZADA) —
    misma fuente que el listado: cotizaciones.referencias_opcion_a."""
    return cotizaciones_service.referencias_opcion_a(db, ids)


def _desgloses_ganadores(db: Session, ids: list[int]) -> dict[int, tuple[Decimal, Decimal]]:
    """(total_mxn, total_usd) de la opción GANADORA de cada CONFIRMADA (F8c):
    el desglose original antes de consolidar con el TC."""
    if not ids:
        return {}
    filas = db.execute(
        select(Solicitud.id, CotizacionOpcion.total_mxn, CotizacionOpcion.total_usd)
        .join(CotizacionOpcion, Solicitud.opcion_seleccionada_id == CotizacionOpcion.id)
        .where(Solicitud.id.in_(ids))
    ).all()
    return {sid: (mxn, usd) for sid, mxn, usd in filas}


def _motivos_rechazo(db: Session, ids: list[int]) -> dict[int, str]:
    """Texto del ÚLTIMO rechazo por solicitud (para filas en RECHAZADA)."""
    if not ids:
        return {}
    filas = db.execute(
        select(HistorialEstado.solicitud_id, MotivoRechazo.texto)
        .join(MotivoRechazo, HistorialEstado.motivo_id == MotivoRechazo.id)
        .where(
            HistorialEstado.solicitud_id.in_(ids),
            HistorialEstado.a == Estado.RECHAZADA,
            or_(HistorialEstado.de.is_(None), HistorialEstado.de != HistorialEstado.a),
        )
        .order_by(HistorialEstado.timestamp, HistorialEstado.id)
    ).tuples()
    return dict(filas)  # el último pisa a los anteriores


def _local(instante: datetime | None, tz: str) -> datetime | None:
    if instante is None:
        return None
    return instante.astimezone(ZoneInfo(tz)).replace(tzinfo=None)


@router.get("/solicitudes/export")
def exportar_solicitudes(
    estado: Estado | None = None,
    prioridad: Prioridad | None = None,
    es_proyecto: bool | None = None,
    cambio_pendiente: bool | None = None,
    # F12 p.5: mismo filtro que el listado (solo filtra para el área compras);
    # el archivo NO lleva columna de fincado en esta versión (llega a Ventas).
    fincada: bool | None = None,
    cliente_id: int | None = None,
    sucursal_id: int | None = None,
    comprador_id: int | None = None,
    vendedor_id: int | None = None,
    desde: date | None = None,
    hasta: date | None = None,
    buscar: str | None = None,
    user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    stmt = solicitudes_service.stmt_listado(
        user,
        estado=estado,
        prioridad=prioridad,
        es_proyecto=es_proyecto,
        cambio_pendiente=cambio_pendiente,
        fincada=fincada,
        cliente_id=cliente_id,
        sucursal_id=sucursal_id,
        comprador_id=comprador_id,
        vendedor_id=vendedor_id,
        desde=desde,
        hasta=hasta,
        buscar=buscar,
    )
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    if total > EXPORT_MAX_FILAS:
        raise AppError(
            422,
            f"El export está limitado a {EXPORT_MAX_FILAS} filas y el filtro "
            f"actual produce {total}; acota el periodo o los filtros",
            "export_demasiado_grande",
        )
    filas = db.execute(stmt.order_by(Solicitud.creado_en.desc(), Solicitud.id.desc())).all()
    solicitudes: list[Solicitud] = [fila[0] for fila in filas]

    ids = [s.id for s in solicitudes]
    sucursales = _sucursales(db, {s.sucursal_id for s in solicitudes})
    usuarios = _nombres_usuarios(
        db,
        {s.vendedor_id for s in solicitudes}
        | {s.comprador_id for s in solicitudes if s.comprador_id is not None},
    )
    ciclos = cargar_ciclos(db, ids)
    tiempos = cargar_tiempos(db, ids)
    referencias = _opciones_a(db, [s.id for s in solicitudes if s.estado == Estado.COTIZADA])
    desgloses = _desgloses_ganadores(
        db, [s.id for s in solicitudes if s.estado == Estado.CONFIRMADA]
    )
    motivos = _motivos_rechazo(db, [s.id for s in solicitudes if s.estado == Estado.RECHAZADA])

    # F14 §0a: el vendedor NO lleva TC ni consolidado (§4.9) — se filtran
    # encabezado Y celda por posición, para no desalinear el resto.
    encabezados = [
        (i, titulo, ancho)
        for i, (titulo, ancho) in enumerate(_ENCABEZADOS)
        if not (user.rol == Rol.VENDEDOR and titulo in _COLUMNAS_SOLO_CONSOLIDADO)
    ]
    posiciones = [i for i, _, _ in encabezados]

    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes"
    ws.append([titulo for _, titulo, _ in encabezados])
    for indice, (_, _, ancho) in enumerate(encabezados, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = ancho

    for solicitud, cliente_nombre in filas:
        sucursal_nombre, tz = sucursales[solicitud.sucursal_id]
        lista_ciclos = ciclos.get(solicitud.id, [])
        ultimo = lista_ciclos[-1] if lista_ciclos else None
        # F8c: desglose por moneda (referencia u opción ganadora), TC y el
        # consolidado MXN de las confirmadas.
        monto_mxn: Decimal | None = None
        monto_usd: Decimal | None = None
        confirmado: Decimal | None = None
        if solicitud.estado == Estado.CONFIRMADA:
            monto_mxn, monto_usd = desgloses.get(solicitud.id, (None, None))
            confirmado = solicitud.monto_confirmado
        elif solicitud.estado == Estado.COTIZADA:
            monto_mxn, monto_usd = referencias.get(solicitud.id, (None, None))
        motivo: str | None = None
        if solicitud.estado == Estado.RECHAZADA:
            motivo = motivos.get(solicitud.id)
        elif solicitud.estado == Estado.NO_CONFIRMADA:
            motivo = solicitud.motivo_no_confirmada

        t = tiempos.get(solicitud.id)
        valores = [
            solicitud.folio,
            cliente_nombre,
            sucursal_nombre,
            usuarios.get(solicitud.vendedor_id),
            usuarios.get(solicitud.comprador_id) if solicitud.comprador_id else None,
            solicitud.estado.value,
            solicitud.prioridad.value,
            "Sí" if solicitud.es_proyecto else "No",
            "Sí" if solicitud.cambio_pendiente else "No",
            _local(solicitud.creado_en, tz),
            _local(solicitud.enviado_en, tz),
            _local(solicitud.cotizado_en, tz),
            _local(solicitud.confirmado_en, tz),
            ultimo.banda.value if ultimo else None,
            round(ultimo.horas_habiles, 2) if ultimo else None,
            t.general_horas_naturales if t else None,
            t.compras_horas_habiles if t else None,
            t.ventas_horas_habiles if t else None,
            monto_mxn or None,
            monto_usd or None,
            solicitud.tipo_cambio,
            confirmado,
            motivo,
        ]
        ws.append([valores[i] for i in posiciones])
        fila_num = ws.max_row
        for col in (10, 11, 12, 13):  # columnas de fecha (corridas por Cambio pendiente)
            ws.cell(row=fila_num, column=col).number_format = _FORMATO_FECHA

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"solicitudes_{datetime.now(UTC).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ------------------------------------------------- export de comparativas (F14 p.4)

_TITULOS_COMPARATIVA: dict[str, str] = {
    "sucursal": "Por sucursal",
    "vendedor": "Por vendedor",
    "cliente": "Por cliente",
    "comprador": "Por comprador",
}

# Formatos de celda con TIPO real (F14 p.4): nada viaja como texto — los
# importes suman y las dinámicas funcionan sin limpiar.
_FMT_MONEDA = "$#,##0.00"
_FMT_HORAS = "0.00"
_FMT_PCT = "0.0%"


def _confirmado_mxn(g: GrupoOut) -> Decimal | None:
    """Mismo criterio que la celda de la UI: sin serie MXN o en cero → vacío
    (None), que el orden manda SIEMPRE al final."""
    valor = g.dinero_confirmado.get("MXN")
    return valor if valor else None


def _rojas(g: GrupoOut) -> int:
    return g.distribucion_bandas.get(Banda.LENTA.value, 0)


# clave ordenable → extractor (espejo de ordenarComparativa del frontend).
_CLAVES_ORDEN: dict[str, Any] = {
    "nombre": lambda g: g.nombre.lower(),
    "volumen": lambda g: g.volumen,
    "ciclos_cerrados": lambda g: g.ciclos_cerrados,
    "mediana_horas_habiles": lambda g: g.mediana_horas_habiles,
    "pct_banda_esperada": lambda g: g.pct_banda_esperada,
    "confirmado_mxn": _confirmado_mxn,
    "rojas": _rojas,
    "carga_abierta": lambda g: g.carga_abierta,
    "cotizadas": lambda g: g.cotizadas,
    "confirmadas": lambda g: g.confirmadas,
    "ratio_confirmacion": lambda g: g.ratio_confirmacion,
}


def _ordenar_comparativa(grupos: list[GrupoOut], orden: str, direccion: str) -> list[GrupoOut]:
    """Orden NUMÉRICO sobre TODO el conjunto, vacíos (None) SIEMPRE al final
    en ambas direcciones; desempate estable por nombre."""
    extractor = _CLAVES_ORDEN[orden]
    con_valor = [g for g in grupos if extractor(g) is not None]
    sin_valor = [g for g in grupos if extractor(g) is None]
    con_valor.sort(key=lambda g: g.nombre.lower())
    con_valor.sort(key=extractor, reverse=direccion == "desc")
    sin_valor.sort(key=lambda g: g.nombre.lower())
    return con_valor + sin_valor


@router.get("/metricas/export-comparativas")
def exportar_comparativas(
    dimension: Dimension,
    orden: str = "volumen",
    direccion: str = "desc",
    desde: date | None = None,
    hasta: date | None = None,
    sucursal_id: int | None = None,
    comprador_id: int | None = None,
    vendedor_id: int | None = None,
    cliente_id: int | None = None,
    prioridad: Prioridad | None = None,
    moneda: Moneda | None = None,
    user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export a Excel de UNA sub-pestaña de Comparativas (F14 p.4), con el
    periodo, filtros y ORDEN vigentes. Gates por dimensión = los de las tablas
    (§2); el rol VENDEDOR hereda el §0: sin columna de dinero consolidado."""
    if dimension == "vendedor" and user.rol not in (
        Rol.ADMIN,
        Rol.DIRECTOR_VENTAS,
        Rol.GERENTE_SUCURSAL,
    ):
        raise AppError(403, "Sin acceso a métricas por vendedor", "sin_permiso")
    if dimension == "comprador" and user.rol not in (Rol.ADMIN, Rol.GERENTE_COMPRAS):
        raise AppError(403, "Sin acceso a métricas por comprador", "sin_permiso")
    if orden not in _CLAVES_ORDEN:
        raise AppError(422, f"Orden desconocido: {orden}", "orden_invalido")
    if direccion not in ("asc", "desc"):
        raise AppError(422, "direccion debe ser asc o desc", "orden_invalido")

    f = Filtros(
        desde=desde,
        hasta=hasta,
        sucursal_id=sucursal_id,
        comprador_id=comprador_id,
        vendedor_id=vendedor_id,
        cliente_id=cliente_id,
        prioridad=prioridad,
        moneda=moneda,
    )
    grupos = _ordenar_comparativa(
        metricas_service.tabla_por(db, user, f, dimension), orden, direccion
    )

    # F14 §0 heredado (§4.9): el VENDEDOR no lleva dinero consolidado — la
    # columna no existe (ni encabezado ni celda). Gerentes, director,
    # comprador y admin sí la llevan.
    con_dinero = user.rol != Rol.VENDEDOR
    titulo = _TITULOS_COMPARATIVA[dimension]
    generado = datetime.now(UTC)

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    ws.append(["Comparativas — Sistema de Cotizaciones Herinox"])
    ws.append(["Pestaña", titulo])
    ws.append(
        [
            "Periodo",
            f"{desde.isoformat() if desde else '(sin desde)'} a "
            f"{hasta.isoformat() if hasta else '(sin hasta)'}",
        ]
    )
    ws.append(["Generado", generado.strftime("%Y-%m-%d %H:%M UTC")])
    ws.append([])

    encabezados = [
        "Nombre",
        "Volumen",
        "Ciclos cerrados",
        "Mediana (h hábiles)",
        "% banda esperada",
        "Verdes",
        "Amarillas",
        "Rojas",
    ]
    if con_dinero:
        encabezados.append("Confirmado (MXN)")
    if dimension == "comprador":
        encabezados.append("Carga abierta")
    if dimension == "cliente":
        encabezados += [
            "Cotizadas",
            "Confirmadas",
            "No confirmadas",
            "Sin desenlace",
            "Ratio confirmación",
        ]
    ws.append(encabezados)

    for g in grupos:
        valores: list[Any] = [
            g.nombre,
            g.volumen,
            g.ciclos_cerrados,
            g.mediana_horas_habiles,
            g.pct_banda_esperada,
            g.distribucion_bandas.get(Banda.ESPERADA.value, 0),
            g.distribucion_bandas.get(Banda.NORMAL.value, 0),
            _rojas(g),
        ]
        formatos: dict[int, str] = {4: _FMT_HORAS, 5: _FMT_PCT}
        if con_dinero:
            valores.append(_confirmado_mxn(g))
            formatos[len(valores)] = _FMT_MONEDA
        if dimension == "comprador":
            valores.append(g.carga_abierta)
        if dimension == "cliente":
            valores += [g.cotizadas, g.confirmadas, g.no_confirmadas, g.sin_desenlace]
            valores.append(g.ratio_confirmacion)
            formatos[len(valores)] = _FMT_PCT
        ws.append(valores)
        fila = ws.max_row
        for col, formato in formatos.items():
            ws.cell(row=fila, column=col).number_format = formato

    for indice, titulo_col in enumerate(encabezados, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = max(14, len(titulo_col) + 4)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    mes = (desde or generado.date()).strftime("%Y-%m")
    nombre = (
        f"Comparativas_{titulo.replace(' ', '-')}_{mes}_{generado.strftime('%Y%m%d-%H%M')}.xlsx"
    )
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
